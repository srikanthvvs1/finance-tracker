"""
sync_drive_to_local.py
──────────────────────
One-way sync: Google Drive + Sheets  →  local documents/ + data.xlsx

Usage:  python sync_drive_to_local.py [--dry-run]

Drive is the golden source. This script:
  1. Exports all Google Sheets data → data.xlsx
  2. Downloads all documents from Drive → documents/<category>/<year>/
  Existing local files are overwritten only if Drive version is newer.
"""

import argparse
import io
import os
import sys
from datetime import datetime
from pathlib import Path

# ─── Config (reuse server_gsheets.py setup) ─────────────────
BASE_DIR = Path(__file__).resolve().parent
CONFIG_DIR = BASE_DIR / "config"

_env_file = CONFIG_DIR / "gsheets.env"
if _env_file.exists():
    for line in _env_file.read_text().splitlines():
        line = line.strip()
        if line and not line.startswith("#") and "=" in line:
            key, val = line.split("=", 1)
            os.environ.setdefault(key.strip(), val.strip())

# ─── Google Auth ─────────────────────────────────────────────
from google.oauth2.service_account import Credentials as SACredentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request as AuthRequest
import pickle
import gspread
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from openpyxl import Workbook

SA_CREDS_FILE = CONFIG_DIR / os.environ.get("GSHEETS_CREDS_FILE", "credentials.json")
SPREADSHEET_ID = os.environ.get("GSHEETS_SPREADSHEET_ID", "")
OAUTH_CLIENT_FILE = CONFIG_DIR / os.environ.get("DRIVE_OAUTH_CLIENT_FILE", "client_secret.json")
OAUTH_TOKEN_FILE = CONFIG_DIR / "drive_token.pickle"

SA_SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]
DRIVE_SCOPES = ["https://www.googleapis.com/auth/drive"]

DRIVE_PARENT_FOLDER_NAME = "Finances"
DRIVE_DOCS_FOLDER_NAME = "FinTrack_Documents"

DOCS_DIR = BASE_DIR / "documents"
XLSX_FILE = BASE_DIR / "data.xlsx"

# Worksheet names and columns (must match server.py / server_gsheets.py)
WORKSHEETS = {
    "Expenses":       ["id", "date", "description", "category", "payment", "amount"],
    "Investments":    ["id", "asset", "name", "category", "units", "buyPrice",
                       "currentPrice", "date", "marketCap", "riskLevel", "ticker", "schemeCode"],
    "Transactions":   ["investmentId", "date", "action", "units", "price"],
    "SavingsGoals":   ["id", "name", "icon", "target", "current", "deadline"],
    "EmergencyFund":  ["target"],
    "EFContributions":["id", "date", "amount", "note"],
    "SavingsHistory": ["month", "income", "expenses", "invested", "emergency", "net_saved"],
}


def get_sheets_client():
    """Get gspread client using service account."""
    creds = SACredentials.from_service_account_file(str(SA_CREDS_FILE), scopes=SA_SCOPES)
    return gspread.authorize(creds)


def get_drive():
    """Get Drive API client using OAuth2 user credentials."""
    creds = None
    if OAUTH_TOKEN_FILE.exists():
        with open(OAUTH_TOKEN_FILE, "rb") as f:
            creds = pickle.load(f)

    if creds and creds.expired and creds.refresh_token:
        creds.refresh(AuthRequest())
    elif not creds or not creds.valid:
        if not OAUTH_CLIENT_FILE.exists():
            # Fall back to SA for read-only operations (listing, downloading)
            print("  ⚠️  No OAuth token — using service account for Drive (read-only)")
            sa_creds = SACredentials.from_service_account_file(str(SA_CREDS_FILE), scopes=SA_SCOPES)
            return build("drive", "v3", credentials=sa_creds, cache_discovery=False)
        flow = InstalledAppFlow.from_client_secrets_file(str(OAUTH_CLIENT_FILE), scopes=DRIVE_SCOPES)
        creds = flow.run_local_server(port=0)
        with open(OAUTH_TOKEN_FILE, "wb") as f:
            pickle.dump(creds, f)

    return build("drive", "v3", credentials=creds, cache_discovery=False)


# ═══════════════════════════════════════════════════════════════
#  1. SYNC SHEETS → data.xlsx
# ═══════════════════════════════════════════════════════════════

def sync_sheets(dry_run=False):
    """Export all Google Sheets worksheets to data.xlsx."""
    print("\n  ━━━ Syncing Google Sheets → data.xlsx ━━━\n")

    gc = get_sheets_client()
    spreadsheet = gc.open_by_key(SPREADSHEET_ID)

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    total_rows = 0
    for ws_name, cols in WORKSHEETS.items():
        try:
            ws = spreadsheet.worksheet(ws_name)
            all_values = ws.get_all_values()
        except gspread.exceptions.WorksheetNotFound:
            print(f"  ⏭️  Worksheet '{ws_name}' not found — skipping")
            continue

        xl_ws = wb.create_sheet(ws_name)
        # Write header
        for c, col_name in enumerate(cols, 1):
            xl_ws.cell(1, c, col_name)

        # Write data (skip header row from Sheets)
        rows = all_values[1:] if len(all_values) > 1 else []
        for r, row in enumerate(rows, 2):
            for c, val in enumerate(row[:len(cols)], 1):
                # Try to preserve numbers
                try:
                    val = float(val)
                    if val == int(val):
                        val = int(val)
                except (ValueError, TypeError):
                    pass
                xl_ws.cell(r, c, val)
        total_rows += len(rows)
        print(f"  ✅ {ws_name}: {len(rows)} rows")

    if dry_run:
        print(f"\n  [DRY RUN] Would write {total_rows} total rows to data.xlsx")
    else:
        wb.save(str(XLSX_FILE))
        print(f"\n  📁 Saved data.xlsx ({total_rows} total rows)")


# ═══════════════════════════════════════════════════════════════
#  2. SYNC DRIVE DOCUMENTS → local documents/
# ═══════════════════════════════════════════════════════════════

def find_folder(drive, name, parent_id):
    """Find a folder by name under parent."""
    q = f"name='{name}' and mimeType='application/vnd.google-apps.folder' and trashed=false"
    if parent_id:
        q += f" and '{parent_id}' in parents"
    results = drive.files().list(q=q, fields="files(id)", pageSize=1).execute()
    files = results.get("files", [])
    return files[0]["id"] if files else None


def sync_documents(dry_run=False):
    """Download all documents from Drive to local documents/ folder."""
    print("\n  ━━━ Syncing Drive Documents → documents/ ━━━\n")

    drive = get_drive()

    # Find Finances / FinTrack_Documents
    fin_id = find_folder(drive, DRIVE_PARENT_FOLDER_NAME, None)
    if not fin_id:
        print("  ⚠️  Finances folder not found on Drive — skipping documents sync")
        return

    docs_id = find_folder(drive, DRIVE_DOCS_FOLDER_NAME, fin_id)
    if not docs_id:
        print("  ⚠️  FinTrack_Documents folder not found — skipping documents sync")
        return

    # List category folders
    q = f"'{docs_id}' in parents and mimeType='application/vnd.google-apps.folder' and trashed=false"
    cats = drive.files().list(q=q, fields="files(id,name)").execute().get("files", [])

    downloaded = 0
    skipped = 0
    total_files = 0

    for cat_folder in sorted(cats, key=lambda f: f["name"]):
        cat_name = cat_folder["name"]
        cat_id = cat_folder["id"]

        # List year subfolders
        yq = f"'{cat_id}' in parents and mimeType='application/vnd.google-apps.folder' and trashed=false"
        years = drive.files().list(q=yq, fields="files(id,name)").execute().get("files", [])

        for yr_folder in sorted(years, key=lambda f: f["name"]):
            year = yr_folder["name"]
            yr_id = yr_folder["id"]

            # List files in year folder
            fq = f"'{yr_id}' in parents and trashed=false and mimeType!='application/vnd.google-apps.folder'"
            files = drive.files().list(
                q=fq, fields="files(id,name,modifiedTime,size)"
            ).execute().get("files", [])

            if not files:
                continue

            local_dir = DOCS_DIR / cat_name / year
            if not dry_run:
                local_dir.mkdir(parents=True, exist_ok=True)

            for f in files:
                total_files += 1
                local_path = local_dir / f["name"]
                drive_modified = f.get("modifiedTime", "")
                drive_size = int(f.get("size", 0))

                # Skip if local file exists and is same size
                if local_path.exists() and local_path.stat().st_size == drive_size:
                    skipped += 1
                    continue

                if dry_run:
                    action = "UPDATE" if local_path.exists() else "NEW"
                    print(f"  [{action}] {cat_name}/{year}/{f['name']} ({drive_size:,} bytes)")
                    downloaded += 1
                    continue

                # Download
                req = drive.files().get_media(fileId=f["id"])
                buf = io.BytesIO()
                downloader = MediaIoBaseDownload(buf, req)
                done = False
                while not done:
                    _, done = downloader.next_chunk()

                buf.seek(0)
                local_path.write_bytes(buf.read())
                downloaded += 1
                print(f"  ✅ {cat_name}/{year}/{f['name']}")

    prefix = "[DRY RUN] Would download" if dry_run else "Downloaded"
    print(f"\n  {prefix}: {downloaded}, Skipped (unchanged): {skipped}, Total on Drive: {total_files}")


# ═══════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="Sync Drive + Sheets → local")
    parser.add_argument("--dry-run", action="store_true", help="Show what would change without writing")
    parser.add_argument("--sheets-only", action="store_true", help="Only sync Sheets → data.xlsx")
    parser.add_argument("--docs-only", action="store_true", help="Only sync Drive documents → documents/")
    args = parser.parse_args()

    print(f"\n  ╔══════════════════════════════════════╗")
    print(f"  ║  FinTrack Sync: Drive → Local        ║")
    print(f"  ║  {datetime.now():%Y-%m-%d %H:%M}                       ║")
    print(f"  ╚══════════════════════════════════════╝")

    if args.dry_run:
        print("\n  🔍 DRY RUN MODE — no files will be written\n")

    do_all = not args.sheets_only and not args.docs_only

    if do_all or args.sheets_only:
        sync_sheets(dry_run=args.dry_run)

    if do_all or args.docs_only:
        sync_documents(dry_run=args.dry_run)

    print("\n  ✨ Sync complete!\n")


if __name__ == "__main__":
    main()
