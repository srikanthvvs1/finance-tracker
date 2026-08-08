"""
FinTrack Local Server
─────────────────────
A lightweight Flask server that:
  1. Serves the static front-end (index.html, script.js, style.css)
  2. Persists all data to an Excel file (data.xlsx) via openpyxl
  3. Proxies live price requests to Yahoo Finance & mfapi.in (no CORS issues)
  4. Imports new expense entries from a Google Sheets inbox at startup

Run:  python server.py
Open: http://localhost:5000
"""

import hashlib
import calendar
import math
import os
import re
import shutil
import sqlite3
import tempfile
import threading
import time
from copy import copy
from datetime import date, datetime, timedelta
from pathlib import Path

from flask import Flask, jsonify, request, send_from_directory
import openpyxl
import requests
import gspread
from gspread.utils import ValueRenderOption
from google.oauth2.service_account import Credentials

# ─── Config (override via environment variables) ────────────────────
BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = Path(os.environ.get("FINTRACK_DATA_DIR", str(BASE_DIR)))
DATA_FILE = DATA_DIR / "data.xlsx"
BACKUP_DIR = DATA_DIR / "backups"
CACHE_DIR = DATA_DIR / "cache"
MARKET_CACHE_FILE = CACHE_DIR / "market_data.sqlite"
BACKUP_KEEP = max(3, int(os.environ.get("FINTRACK_BACKUP_KEEP", "20")))
_file_lock = threading.Lock()
_market_cache_lock = threading.Lock()

CONFIG_DIR = BASE_DIR / "config"
_env_file = CONFIG_DIR / "gsheets.env"
if _env_file.exists():
    for line in _env_file.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if line and not line.startswith("#") and "=" in line:
            key, value = line.split("=", 1)
            os.environ.setdefault(key.strip(), value.strip())

GSHEETS_CREDS_FILE = CONFIG_DIR / os.environ.get("GSHEETS_CREDS_FILE", "credentials.json")
GSHEETS_SPREADSHEET_ID = os.environ.get("GSHEETS_SPREADSHEET_ID", "")
FORM_SHEET_NAME = "FormExpenses"
GSHEETS_SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
]
_sync_lock = threading.Lock()
_last_sync_result = {"synced": 0, "errors": [], "timestamp": None}

STATIC_DIR = BASE_DIR / "static"
app = Flask(__name__, static_folder=str(STATIC_DIR), static_url_path="")


# ═══════════════════════════════════════════════════════════════
#  EXCEL HELPERS  (atomic writes via temp-file + rename)
# ═══════════════════════════════════════════════════════════════

EXPENSE_COLS = [
    "id", "date", "description", "category", "payment", "amount", "accountId",
    "expenseNature",
]
INCOME_COLS = ["id", "date", "source", "description", "amount", "accountId"]
INVESTMENT_COLS = [
    "id", "asset", "name", "category", "units", "buyPrice",
    "currentPrice", "date", "marketCap", "riskLevel", "ticker", "schemeCode",
    "containerAccountId", "entryMode",
]
TRANSACTION_COLS = [
    "investmentId", "date", "action", "units", "price", "accountId", "source",
    "settlementDate", "charges",
]
SAVINGS_HIST_COLS = [
    "month", "income", "expenses", "invested", "emergency", "net_saved", "accountId",
]
SAVINGS_GOAL_COLS = ["id", "name", "icon", "target", "current", "deadline"]
EMERGENCY_COLS = ["target"]
EMERGENCY_CONTRIB_COLS = ["id", "date", "amount", "note"]
EMERGENCY_ALLOCATION_COLS = [
    "id", "sourceType", "sourceId", "allocationMode", "amount",
    "liquidity", "note", "updatedAt",
]
BUDGET_COLS = ["month", "category", "amount"]
RECURRING_BILL_COLS = [
    "id", "name", "category", "amount", "dueDay", "frequency", "active",
    "includedInBudget",
]
NET_WORTH_COLS = [
    "month", "cash", "bank", "investments", "retirement", "otherAssets",
    "loans", "creditCards", "otherLiabilities",
]
NET_WORTH_AUTO_COLS = [
    "month", "asOf", "cash", "bank", "investments", "retirement",
    "otherAssets", "loans", "creditCards", "otherLiabilities",
]
CASH_FLOW_COLS = ["month", "openingBalance", "otherIncome", "safetyBalance"]
ACCOUNT_COLS = [
    "id", "name", "bank", "type", "classification", "purpose", "currency", "openingDate",
    "openingBalance", "statementBalance", "creditLimit", "includeNetWorth", "active",
    "settlementAccountId",
]
TRANSFER_COLS = ["id", "date", "fromAccountId", "toAccountId", "amount", "note"]
RECONCILIATION_ADJUSTMENT_COLS = [
    "id", "accountId", "date", "amount", "reason", "createdAt",
]
RECURRING_RULE_COLS = [
    "id", "name", "type", "frequency", "day", "amount", "fromAccountId",
    "investmentId", "startDate", "endDate", "active",
]
RECURRING_OCCURRENCE_COLS = [
    "id", "ruleId", "scheduledDate", "status", "actualDate", "actualAmount",
    "units", "price", "note",
]

PLANNING_SHEETS = {
    "IncomeTransactions": INCOME_COLS,
    "Budgets": BUDGET_COLS,
    "RecurringBills": RECURRING_BILL_COLS,
    "NetWorth": NET_WORTH_COLS,
    "NetWorthAuto": NET_WORTH_AUTO_COLS,
    "CashFlow": CASH_FLOW_COLS,
    "Accounts": ACCOUNT_COLS,
    "Transfers": TRANSFER_COLS,
    "ReconciliationAdjustments": RECONCILIATION_ADJUSTMENT_COLS,
    "RecurringRules": RECURRING_RULE_COLS,
    "RecurringOccurrences": RECURRING_OCCURRENCE_COLS,
    "EmergencyAllocations": EMERGENCY_ALLOCATION_COLS,
}


def _migrate_legacy_income(wb):
    """Create dated income entries from legacy monthly totals exactly once."""
    if "IncomeTransactions" not in wb.sheetnames or "SavingsHistory" not in wb.sheetnames:
        return False
    income_ws = wb["IncomeTransactions"]
    if income_ws.max_row > 1:
        return False

    savings_ws = wb["SavingsHistory"]
    headers = {str(cell.value): index for index, cell in enumerate(savings_ws[1]) if cell.value}
    required = {"month", "income", "accountId"}
    if not required.issubset(headers):
        return False

    additions = []
    for row in savings_ws.iter_rows(min_row=2, values_only=True):
        label = row[headers["month"]] if headers["month"] < len(row) else None
        amount = row[headers["income"]] if headers["income"] < len(row) else None
        account_id = row[headers["accountId"]] if headers["accountId"] < len(row) else None
        parsed = _parse_month_label(label)
        try:
            amount = float(amount or 0)
        except (TypeError, ValueError):
            amount = 0
        if not parsed or amount <= 0 or not account_id:
            continue
        year, month = parsed
        paid_date = date(year, month, calendar.monthrange(year, month)[1])
        additions.append([
            len(additions) + 1,
            paid_date,
            "salary",
            f"Migrated monthly income - {label}",
            amount,
            account_id,
        ])

    for item in additions:
        income_ws.append(item)
        income_ws.cell(row=income_ws.max_row, column=2).number_format = DATE_FMT
    return bool(additions)


def _ensure_planning_sheets(wb):
    """Add newly introduced planning sheets without touching existing user data."""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    changed = False
    for sheet_name, columns in PLANNING_SHEETS.items():
        if sheet_name in wb.sheetnames:
            continue
        ws = wb.create_sheet(sheet_name)
        ws.append(columns)
        for cell in ws[1]:
            cell.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"),
            )
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        if sheet_name == "EmergencyAllocations":
            widths = [14, 16, 16, 18, 16, 18, 36, 24]
            for column_index, width in enumerate(widths, 1):
                letter = openpyxl.utils.get_column_letter(column_index)
                ws.column_dimensions[letter].width = width
        changed = True
    return changed


_VALID_EXPENSE_NATURES = {"fixed", "variable"}
_DEFAULT_FIXED_CATEGORIES = {"housing", "subscriptions"}
_FIXED_EXPENSE_PATTERNS = (
    r"\brent\b", r"\blease\b", r"\bemi\b", r"\bmortgage\b",
    r"\bsubscription\b", r"\bmembership\b", r"\binsurance\b",
    r"\bschool fees?\b", r"\btuition fees?\b", r"\bannual charges?\b",
    r"\bbroadband\b", r"\binternet\b", r"\bwifi\b",
)


def _infer_expense_nature(category, description=""):
    """Infer a sensible editable default without changing the expense category."""
    if str(category or "").strip().lower() in _DEFAULT_FIXED_CATEGORIES:
        return "fixed"
    text = str(description or "").strip().casefold()
    return "fixed" if any(re.search(pattern, text) for pattern in _FIXED_EXPENSE_PATTERNS) else "variable"


def _normalise_expense_nature(raw, category, description=""):
    cleaned = str(raw or "").strip().lower()
    if cleaned in _VALID_EXPENSE_NATURES:
        return cleaned
    return _infer_expense_nature(category, description)


def _ensure_expense_nature_column(wb):
    """Append/backfill the additive nature column while preserving all existing cells."""
    if "Expenses" not in wb.sheetnames:
        return False
    ws = wb["Expenses"]
    headers = [str(cell.value or "") for cell in ws[1]]
    positions = {name: index + 1 for index, name in enumerate(headers) if name}
    changed = False
    nature_col = positions.get("expenseNature")
    if nature_col is None:
        nature_col = len(headers) + 1
        source = ws.cell(row=1, column=max(1, nature_col - 1))
        target = ws.cell(row=1, column=nature_col, value="expenseNature")
        target._style = copy(source._style)
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)
        ws.column_dimensions[openpyxl.utils.get_column_letter(nature_col)].width = 18
        changed = True

    category_col = positions.get("category")
    description_col = positions.get("description")
    for row_number in range(2, ws.max_row + 1):
        category = ws.cell(row=row_number, column=category_col).value if category_col else ""
        description = ws.cell(row=row_number, column=description_col).value if description_col else ""
        cell = ws.cell(row=row_number, column=nature_col)
        normalised = _normalise_expense_nature(cell.value, category, description)
        if cell.value != normalised:
            cell.value = normalised
            changed = True

    if changed:
        ws.auto_filter.ref = ws.dimensions
    return changed


_VEGETABLES_FRUITS_DESCRIPTION = re.compile(
    r"^\s*(?:vegetables?|veggies?|fruits?|produce)(?:\s*[-:/]|\s*$)",
    re.IGNORECASE,
)


def _migrate_vegetables_fruits_category(wb):
    """Move only unmistakable legacy Grocery rows into the produce category."""
    if "Expenses" not in wb.sheetnames:
        return False
    ws = wb["Expenses"]
    positions = {
        str(cell.value or ""): index + 1
        for index, cell in enumerate(ws[1])
        if cell.value
    }
    category_col = positions.get("category")
    description_col = positions.get("description")
    if not category_col or not description_col:
        return False

    changed = False
    for row_number in range(2, ws.max_row + 1):
        category_cell = ws.cell(row=row_number, column=category_col)
        description = ws.cell(row=row_number, column=description_col).value
        if (
            str(category_cell.value or "").strip().lower() == "grocery"
            and _VEGETABLES_FRUITS_DESCRIPTION.search(str(description or ""))
        ):
            category_cell.value = "vegetables_fruits"
            changed = True
    return changed


def _migrate_sheet_schema(wb, sheet_name, columns):
    """Reorder/add known columns without discarding existing rows."""
    if sheet_name not in wb.sheetnames:
        return False
    ws = wb[sheet_name]
    existing = [cell.value for cell in ws[1]]
    if existing == columns:
        return False
    positions = {str(name): index for index, name in enumerate(existing) if name}
    values = list(ws.iter_rows(min_row=2, values_only=True))
    migrated = [
        [row[positions[col]] if col in positions and positions[col] < len(row) else None
         for col in columns]
        for row in values
    ]
    ws.delete_rows(1, ws.max_row)
    ws.append(columns)
    for row in migrated:
        ws.append(row)
    return True


def _safe_save(wb):
    """Save workbook via temp file + rename to avoid corruption."""
    fd, tmp_path = tempfile.mkstemp(suffix=".xlsx", dir=str(DATA_DIR))
    os.close(fd)
    try:
        wb.save(tmp_path)
        wb.close()
        # Retry the atomic replacement in case Excel, OneDrive, or antivirus
        # briefly holds the destination file.
        for i in range(5):
            try:
                os.replace(tmp_path, DATA_FILE)
                return
            except PermissionError:
                time.sleep(0.3 * (i + 1))
        os.replace(tmp_path, DATA_FILE)
    except Exception:
        # Clean up temp file on failure
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
        raise


def _backup_workbook():
    """Create a timestamped rolling backup before modifying data.xlsx."""
    if not DATA_FILE.exists():
        return None
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    destination = BACKUP_DIR / f"data_{stamp}.xlsx"
    shutil.copy2(DATA_FILE, destination)
    backups = sorted(BACKUP_DIR.glob("data_*.xlsx"), key=lambda path: path.stat().st_mtime)
    for old_backup in backups[:-BACKUP_KEEP]:
        old_backup.unlink(missing_ok=True)
    return destination


def _ensure_workbook():
    """Create an empty Excel file with formatted sheets if it does not exist."""
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    if DATA_FILE.exists():
        try:
            wb = _open_workbook(retries=8, delay=1.0)
            # Upgrade known sheets in place; never delete user data for a schema change.
            changed = _ensure_planning_sheets(wb)
            changed = _ensure_expense_nature_column(wb) or changed
            changed = _migrate_vegetables_fruits_category(wb) or changed
            schemas = {
                "Expenses": EXPENSE_COLS,
                "Investments": INVESTMENT_COLS,
                "Transactions": TRANSACTION_COLS,
                "SavingsGoals": SAVINGS_GOAL_COLS,
                "EmergencyFund": EMERGENCY_COLS,
                "EFContributions": EMERGENCY_CONTRIB_COLS,
                "SavingsHistory": SAVINGS_HIST_COLS,
                **PLANNING_SHEETS,
            }
            for sheet_name, columns in schemas.items():
                changed = _migrate_sheet_schema(wb, sheet_name, columns) or changed
            changed = _migrate_legacy_income(wb) or changed
            if changed:
                _backup_workbook()
                _safe_save(wb)
                print("[OK] Migrated workbook schema without deleting existing data")
            else:
                wb.close()
            return
        except PermissionError:
            print("[!] data.xlsx is locked. Close Excel/OneDrive and restart FinTrack.")
            return
        except Exception as exc:
            BACKUP_DIR.mkdir(parents=True, exist_ok=True)
            stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            quarantined = BACKUP_DIR / f"data_unreadable_{stamp}.xlsx"
            shutil.move(str(DATA_FILE), str(quarantined))
            print(f"[!] Workbook could not be opened ({exc}). Preserved it as {quarantined.name}.")

    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    def _create_sheet(name, cols, widths=None):
        ws = wb.create_sheet(name)
        ws.append(cols)
        for c_idx, col in enumerate(cols, 1):
            cell = ws.cell(row=1, column=c_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        if widths:
            for c_idx, w in enumerate(widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(c_idx)].width = w
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        return ws

    # Create empty, formatted sheets. User data is added through the app or expense sync.
    _create_sheet("Expenses", EXPENSE_COLS, widths=[12, 14, 35, 16, 14, 14, 16, 18])
    _create_sheet("IncomeTransactions", INCOME_COLS, widths=[12, 14, 18, 36, 16, 16])
    _create_sheet("Investments", INVESTMENT_COLS, widths=[14, 14, 28, 18, 12, 16, 16, 14, 12, 12, 16, 14, 18, 14])
    _create_sheet("Transactions", TRANSACTION_COLS, widths=[14, 14, 10, 12, 16, 14, 14, 16, 14])
    _create_sheet("SavingsGoals", SAVINGS_GOAL_COLS, widths=[12, 24, 8, 16, 16, 14])
    _create_sheet("EmergencyFund", EMERGENCY_COLS, widths=[16])
    _create_sheet("EFContributions", EMERGENCY_CONTRIB_COLS, widths=[12, 14, 16, 30])
    _create_sheet(
        "EmergencyAllocations", EMERGENCY_ALLOCATION_COLS,
        widths=[14, 16, 16, 18, 16, 18, 36, 24],
    )
    _create_sheet("SavingsHistory", SAVINGS_HIST_COLS, widths=[14, 16, 16, 16, 16, 16])
    _create_sheet("Budgets", BUDGET_COLS, widths=[14, 18, 16])
    _create_sheet("RecurringBills", RECURRING_BILL_COLS, widths=[12, 28, 18, 16, 12, 14, 12])
    _create_sheet("NetWorth", NET_WORTH_COLS, widths=[14, 16, 16, 16, 16, 16, 16, 16, 18])
    _create_sheet("NetWorthAuto", NET_WORTH_AUTO_COLS, widths=[14, 14, 16, 16, 16, 16, 16, 16, 16, 18])
    _create_sheet("CashFlow", CASH_FLOW_COLS, widths=[14, 18, 16, 18])
    _create_sheet("Accounts", ACCOUNT_COLS, widths=[12, 24, 18, 18, 16, 18, 12, 14, 18, 18, 16, 18, 12, 20])
    _create_sheet("Transfers", TRANSFER_COLS, widths=[12, 14, 18, 18, 16, 30])
    _create_sheet(
        "ReconciliationAdjustments", RECONCILIATION_ADJUSTMENT_COLS,
        widths=[18, 14, 14, 18, 42, 24],
    )
    _create_sheet("RecurringRules", RECURRING_RULE_COLS, widths=[12, 28, 14, 14, 10, 16, 18, 18, 14, 14, 12])
    _create_sheet("RecurringOccurrences", RECURRING_OCCURRENCE_COLS, widths=[24, 12, 16, 14, 16, 18, 14, 16, 32])

    _safe_save(wb)
    print("[OK] Created empty data.xlsx workbook")

# Columns that hold date values (written as Excel dates, read back as ISO strings)
DATE_COLUMNS = {"date", "deadline", "openingDate", "asOf", "settlementDate"}
DATE_FMT = "YYYY-MM-DD"   # Excel custom number format


def _to_date(val):
    """Convert an ISO date string (e.g. '2026-04-01') to a datetime.date for Excel."""
    if isinstance(val, (date, datetime)):
        return val
    if isinstance(val, str) and len(val) >= 10:
        try:
            return datetime.strptime(val[:10], "%Y-%m-%d").date()
        except ValueError:
            pass
    return val


def _from_date(val):
    """Convert a datetime.date/datetime read from Excel back to an ISO string."""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.isoformat()
    return val


def _open_workbook(read_only=False, retries=5, delay=0.5):
    """Open data.xlsx with retries to handle transient OneDrive / antivirus locks."""
    for attempt in range(retries):
        try:
            return openpyxl.load_workbook(str(DATA_FILE), read_only=read_only)
        except PermissionError:
            if attempt < retries - 1:
                time.sleep(delay * (attempt + 1))
            else:
                raise


def _read_sheet(sheet_name, columns):
    """Read an entire sheet into a list of dicts."""
    with _file_lock:
        wb = _open_workbook(read_only=True)
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            obj = {}
            for i, col in enumerate(columns):
                val = row[i] if i < len(row) else None
                if col in DATE_COLUMNS:
                    val = _from_date(val)
                obj[col] = val
            rows.append(obj)
        wb.close()
    return rows


def _write_sheet(sheet_name, columns, rows):
    """Overwrite a single sheet."""
    with _file_lock:
        _backup_workbook()
        wb = _open_workbook()
        ws = wb[sheet_name]
        ws.delete_rows(2, ws.max_row)
        for obj in rows:
            vals = []
            for c in columns:
                v = obj.get(c)
                if c in DATE_COLUMNS:
                    v = _to_date(v)
                vals.append(v)
            ws.append(vals)
        # Apply date format to date columns
        date_col_indices = [i + 1 for i, c in enumerate(columns) if c in DATE_COLUMNS]
        for r in range(2, ws.max_row + 1):
            for ci in date_col_indices:
                ws.cell(row=r, column=ci).number_format = DATE_FMT
        _safe_save(wb)


def _write_sheets(sheet_data):
    """Write multiple sheets atomically in one lock."""
    with _file_lock:
        _backup_workbook()
        wb = _open_workbook()
        for sheet_name, columns, rows in sheet_data:
            ws = wb[sheet_name]
            ws.delete_rows(2, ws.max_row)
            for obj in rows:
                vals = []
                for c in columns:
                    v = obj.get(c)
                    if c in DATE_COLUMNS:
                        v = _to_date(v)
                    vals.append(v)
                ws.append(vals)
            # Apply date format
            date_col_indices = [i + 1 for i, c in enumerate(columns) if c in DATE_COLUMNS]
            for r in range(2, ws.max_row + 1):
                for ci in date_col_indices:
                    ws.cell(row=r, column=ci).number_format = DATE_FMT
        _safe_save(wb)


# ═══════════════════════════════════════════════════════════════
#  GOOGLE SHEETS EXPENSE INBOX → LOCAL EXCEL
# ═══════════════════════════════════════════════════════════════

_VALID_CATEGORIES = {
    "food", "grocery", "vegetables_fruits", "travel", "housing", "parents_fund", "health",
    "personal_care", "subscriptions", "entertainment", "utilities", "shopping", "other",
}
_VALID_PAYMENTS = {"card", "debit", "cash", "transfer", "upi"}


def _normalise_category(raw):
    cleaned = str(raw or "").strip().lower().replace(" ", "_")
    if cleaned in _VALID_CATEGORIES:
        return cleaned
    return {
        "food_&_dining": "food", "dining": "food", "groceries": "grocery",
        "vegetable": "vegetables_fruits", "vegetables": "vegetables_fruits",
        "veggie": "vegetables_fruits", "veggies": "vegetables_fruits",
        "fruit": "vegetables_fruits", "fruits": "vegetables_fruits",
        "produce": "vegetables_fruits", "vegetables_&_fruits": "vegetables_fruits",
        "vegetables_and_fruits": "vegetables_fruits",
        "parent": "parents_fund", "parents": "parents_fund",
        "parent_fund": "parents_fund", "parents_support": "parents_fund",
        "parent_support": "parents_fund", "parental_support": "parents_fund",
        "family_support": "parents_fund", "money_to_parents": "parents_fund",
        "transport": "travel", "transportation": "travel", "cab": "travel",
        "rent": "housing", "home": "housing", "medical": "health",
        "medicine": "health", "pharmacy": "health", "bills": "utilities",
        "haircut": "personal_care", "salon": "personal_care",
        "grooming": "personal_care", "cosmetics": "personal_care",
        "subscription": "subscriptions", "subscriptions_&_software": "subscriptions",
        "software": "subscriptions", "chatgpt": "subscriptions",
        "cloud_storage": "subscriptions", "antivirus": "subscriptions",
        "electricity": "utilities", "internet": "utilities",
        "movies": "entertainment", "games": "entertainment",
        "clothes": "shopping", "amazon": "shopping",
    }.get(cleaned, "other")


def _normalise_payment(raw):
    cleaned = str(raw or "").strip().lower().replace(" ", "_")
    if cleaned in _VALID_PAYMENTS:
        return cleaned
    return {
        "credit_card": "card", "credit": "card", "cc": "card",
        "debit_card": "debit", "dc": "debit", "bank_transfer": "transfer",
        "neft": "transfer", "imps": "transfer", "google_pay": "upi",
        "gpay": "upi", "phonepe": "upi", "paytm": "upi",
    }.get(cleaned, "card")


def _normalise_form_date(raw):
    # Unformatted Google Sheets date cells are numeric serials. Reading the
    # serial avoids ambiguous locale strings such as 8/1/2026 (Aug 1 vs Jan 8).
    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        try:
            return (date(1899, 12, 30) + timedelta(days=int(raw))).isoformat()
        except (OverflowError, ValueError):
            pass
    value = str(raw or "").strip()
    for fmt in (
        "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d",
        "%d %b %Y", "%d %B %Y", "%m/%d/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S",
    ):
        try:
            return datetime.strptime(value, fmt).date().isoformat()
        except ValueError:
            continue
    return date.today().isoformat()


def _expense_fingerprint(expense):
    amount = float(expense.get("amount") or 0)
    raw = "|".join((
        str(expense.get("date") or ""),
        f"{amount:.2f}",
        str(expense.get("description") or "").strip().casefold(),
    ))
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _open_form_worksheet():
    """Connect lazily so local operation does not depend on Google being available."""
    if not GSHEETS_SPREADSHEET_ID:
        raise RuntimeError("GSHEETS_SPREADSHEET_ID is not configured")
    if not GSHEETS_CREDS_FILE.exists():
        raise RuntimeError(f"Google credentials not found: {GSHEETS_CREDS_FILE}")
    credentials = Credentials.from_service_account_file(
        str(GSHEETS_CREDS_FILE), scopes=GSHEETS_SCOPES
    )
    spreadsheet = gspread.authorize(credentials).open_by_key(GSHEETS_SPREADSHEET_ID)
    return spreadsheet.worksheet(FORM_SHEET_NAME)


def _sync_form_expenses():
    """Move valid, unseen expense-inbox rows from Google Sheets into data.xlsx."""
    global _last_sync_result
    with _sync_lock:
        errors = []
        timestamp = datetime.now().isoformat()
        try:
            worksheet = _open_form_worksheet()
            sheet_rows = worksheet.get_all_values(
                value_render_option=ValueRenderOption.unformatted
            )
            if len(sheet_rows) <= 1:
                _last_sync_result = {"synced": 0, "errors": [], "timestamp": timestamp}
                return 0, []

            header = {name.strip(): index for index, name in enumerate(sheet_rows[0])}

            def field(row, *names):
                for name in names:
                    index = header.get(name)
                    if index is not None and index < len(row):
                        return row[index]
                return ""

            local_expenses = _read_sheet("Expenses", EXPENSE_COLS)
            account_rows = _read_sheet("Accounts", ACCOUNT_COLS)
            default_spending_account = next(
                (
                    int(row["id"]) for row in account_rows
                    if row.get("id") and str(row.get("purpose") or "").lower() == "spending"
                    and row.get("active") is not False
                ),
                None,
            )
            if default_spending_account is None:
                raise RuntimeError(
                    "Add an active Spending account before importing form expenses"
                )
            existing = {_expense_fingerprint(item) for item in local_expenses}
            next_id = max((int(item.get("id") or 0) for item in local_expenses), default=0)
            additions = []
            processed_rows = []

            for sheet_row_number, row in enumerate(sheet_rows[1:], start=2):
                try:
                    amount_text = str(field(row, "Amount")).replace(",", "").strip()
                    amount = float(amount_text)
                    if amount <= 0:
                        raise ValueError("amount must be greater than zero")
                    raw_expense_date = field(row, "Expense Date", "ExpenseDate")
                    if not str(raw_expense_date or "").strip():
                        raw_expense_date = field(row, "Timestamp")
                    item = {
                        "date": _normalise_form_date(raw_expense_date),
                        "description": str(field(row, "Description")).strip() or "Form entry",
                        "category": _normalise_category(field(row, "Category")),
                        "payment": _normalise_payment(field(row, "PaymentMode", "Payment Mode")),
                        "amount": amount,
                        "accountId": default_spending_account,
                    }
                    item["expenseNature"] = _normalise_expense_nature(
                        field(row, "Expense Nature", "ExpenseNature", "Nature", "Expense Type"),
                        item["category"],
                        item["description"],
                    )
                    fingerprint = _expense_fingerprint(item)
                    if fingerprint not in existing:
                        next_id += 1
                        item["id"] = next_id
                        additions.append(item)
                        existing.add(fingerprint)
                    processed_rows.append(sheet_row_number)
                except (TypeError, ValueError) as exc:
                    errors.append(f"Row {sheet_row_number}: {exc}")

            if additions:
                _write_sheet("Expenses", EXPENSE_COLS, local_expenses + additions)

            # Clear only successfully imported or duplicate rows. Invalid rows remain visible.
            for row_number in reversed(processed_rows):
                worksheet.delete_rows(row_number)

            _last_sync_result = {
                "synced": len(additions), "errors": errors, "timestamp": timestamp
            }
            return len(additions), errors
        except Exception as exc:
            errors.append(f"Google Sheets sync failed: {exc}")
            _last_sync_result = {"synced": 0, "errors": errors, "timestamp": timestamp}
            return 0, errors

# ═══════════════════════════════════════════════════════════════
#  STATIC FILE SERVING
# ═══════════════════════════════════════════════════════════════

@app.route("/")
def index():
    return send_from_directory(str(STATIC_DIR), "index.html")


# ═══════════════════════════════════════════════════════════════
#  API: USER INFO & CONFIG
# ═══════════════════════════════════════════════════════════════

@app.route("/api/user-info")
def user_info():
    """Return the OS full display name and current data directory."""
    fullname = None
    # Try Windows API for display name (e.g. "John Doe")
    try:
        import ctypes
        GetUserNameExW = ctypes.windll.secur32.GetUserNameExW
        NameDisplay = 3  # EXTENDED_NAME_FORMAT → NameDisplay
        buf = ctypes.create_unicode_buffer(256)
        size = ctypes.pointer(ctypes.c_ulong(256))
        if GetUserNameExW(NameDisplay, buf, size):
            fullname = buf.value
    except Exception:
        pass
    if not fullname:
        fullname = os.environ.get("USERNAME") or os.environ.get("USER") or "User"
    return jsonify({
        "username": fullname,
        "initials": "".join(w[0].upper() for w in fullname.split() if w)[:2] or fullname[:2].upper(),
        "dataDir": str(DATA_DIR),
    })


# ═══════════════════════════════════════════════════════════════
#  API: EXPENSES
# ═══════════════════════════════════════════════════════════════

@app.route("/api/expenses", methods=["GET"])
def get_expenses():
    rows = _read_sheet("Expenses", EXPENSE_COLS)
    # Ensure numeric types
    for r in rows:
        r["id"] = int(r["id"]) if r["id"] else 0
        r["amount"] = float(r["amount"]) if r["amount"] else 0
        r["accountId"] = int(r["accountId"]) if r.get("accountId") else None
        r["expenseNature"] = _normalise_expense_nature(
            r.get("expenseNature"), r.get("category"), r.get("description")
        )
    return jsonify(rows)


@app.route("/api/expenses", methods=["POST"])
def save_expenses():
    data = request.get_json(force=True)
    if not isinstance(data, list):
        return jsonify({"error": "Expected array"}), 400
    accounts = {
        int(row["id"]): row for row in _read_sheet("Accounts", ACCOUNT_COLS) if row.get("id")
    }
    cleaned = []
    for row in data:
        if not isinstance(row, dict):
            return jsonify({"error": "Invalid expense"}), 400
        try:
            amount = float(row.get("amount") or 0)
            account_id = int(row.get("accountId"))
        except (TypeError, ValueError):
            return jsonify({"error": "Each expense requires a paying account and numeric amount"}), 400
        account = accounts.get(account_id)
        if amount <= 0 or not account or not row.get("date"):
            return jsonify({"error": "Expense requires a date, positive amount, and valid paying account"}), 400
        if str(account.get("classification") or "asset").lower() == "investment":
            return jsonify({"error": "Expenses cannot be paid directly from an investment account"}), 400
        raw_nature = str(row.get("expenseNature") or "").strip().lower()
        if raw_nature and raw_nature not in _VALID_EXPENSE_NATURES:
            return jsonify({"error": "Expense nature must be fixed or variable"}), 400
        expense_nature = _normalise_expense_nature(
            raw_nature, row.get("category"), row.get("description")
        )
        cleaned.append({
            **{column: row.get(column) for column in EXPENSE_COLS},
            "amount": amount,
            "accountId": account_id,
            "expenseNature": expense_nature,
        })
    _write_sheet("Expenses", EXPENSE_COLS, cleaned)
    return jsonify({"ok": True, "count": len(cleaned)})


# Income is stored as dated transactions so several sources can credit several accounts.
INCOME_SOURCES = {
    "salary", "bonus", "freelance", "business", "interest", "dividend",
    "rent", "gift", "other",
}


@app.route("/api/income-transactions", methods=["GET"])
def get_income_transactions():
    rows = _read_sheet("IncomeTransactions", INCOME_COLS)
    for row in rows:
        row["id"] = int(row["id"]) if row.get("id") else 0
        row["amount"] = float(row["amount"]) if row.get("amount") else 0
        row["accountId"] = int(row["accountId"]) if row.get("accountId") else None
        row["source"] = str(row.get("source") or "other").lower()
    return jsonify(rows)


@app.route("/api/income-transactions", methods=["POST"])
def save_income_transactions():
    data = request.get_json(force=True)
    if not isinstance(data, list):
        return jsonify({"error": "Expected array"}), 400
    accounts = {
        int(row["id"]): row for row in _read_sheet("Accounts", ACCOUNT_COLS)
        if row.get("id")
    }
    cleaned = []
    seen_ids = set()
    for row in data:
        if not isinstance(row, dict):
            return jsonify({"error": "Invalid income transaction"}), 400
        try:
            row_id = int(row.get("id"))
            amount = float(row.get("amount") or 0)
            account_id = int(row.get("accountId"))
        except (TypeError, ValueError):
            return jsonify({"error": "Income requires an id, amount, and credited account"}), 400
        source = str(row.get("source") or "other").lower()
        account = accounts.get(account_id)
        if (row_id in seen_ids or amount <= 0 or not row.get("date") or not account):
            return jsonify({"error": "Income requires a unique id, date, positive amount, and valid account"}), 400
        if source not in INCOME_SOURCES:
            return jsonify({"error": "Invalid income source"}), 400
        if (str(account.get("classification") or "asset").lower() in {"liability", "investment"}
                or str(account.get("type") or "").lower() in {
                    "credit_card", "loan", "demat", "mutual_fund", "gold", "ppf", "nps", "fixed_deposit",
                }):
            return jsonify({"error": "Income must credit a bank, cash, or wallet asset account"}), 400
        seen_ids.add(row_id)
        cleaned.append({
            "id": row_id,
            "date": str(row.get("date"))[:10],
            "source": source,
            "description": str(row.get("description") or "").strip()[:160],
            "amount": amount,
            "accountId": account_id,
        })
    _write_sheet("IncomeTransactions", INCOME_COLS, cleaned)
    return jsonify({"ok": True, "count": len(cleaned)})


@app.route("/api/sync-form-expenses", methods=["POST"])
def api_sync_form_expenses():
    """Manually pull new expense-inbox rows into the local workbook."""
    synced, errors = _sync_form_expenses()
    status = 200 if not errors or synced else 502
    return jsonify({
        "synced": synced,
        "errors": errors,
        "timestamp": _last_sync_result["timestamp"],
    }), status


@app.route("/api/sync-form-expenses", methods=["GET"])
def api_sync_status():
    return jsonify(_last_sync_result)


# ═══════════════════════════════════════════════════════════════
#  API: INVESTMENTS  (includes nested transactions)
# ═══════════════════════════════════════════════════════════════

@app.route("/api/investments", methods=["GET"])
def get_investments():
    invs = _read_sheet("Investments", INVESTMENT_COLS)
    txns = _read_sheet("Transactions", TRANSACTION_COLS)

    # Ensure numeric types
    for r in invs:
        r["id"] = int(r["id"]) if r["id"] else 0
        r["units"] = float(r["units"]) if r["units"] else 0
        r["buyPrice"] = float(r["buyPrice"]) if r["buyPrice"] else 0
        r["currentPrice"] = float(r["currentPrice"]) if r["currentPrice"] else 0
        r["containerAccountId"] = int(r["containerAccountId"]) if r.get("containerAccountId") else None
        r["entryMode"] = str(r.get("entryMode") or "connected")
        r["transactions"] = []

    inv_map = {r["id"]: r for r in invs}
    for t in txns:
        inv_id = int(t["investmentId"]) if t["investmentId"] else 0
        if inv_id in inv_map:
            inv_map[inv_id]["transactions"].append({
                "date": t["date"],
                "action": t["action"],
                "units": float(t["units"]) if t["units"] else 0,
                "price": float(t["price"]) if t["price"] else 0,
                "accountId": int(t["accountId"]) if t.get("accountId") else None,
                "source": str(t.get("source") or "connected"),
                "settlementDate": str(t.get("settlementDate") or "")[:10],
                "charges": float(t["charges"]) if t.get("charges") else 0,
            })

    return jsonify(invs)


@app.route("/api/investments", methods=["POST"])
def save_investments():
    data = request.get_json(force=True)
    if not isinstance(data, list):
        return jsonify({"error": "Expected array"}), 400

    account_rows = _read_sheet("Accounts", ACCOUNT_COLS)
    account_map = {
        int(row["id"]): row
        for row in account_rows if row.get("id")
    }
    account_types = {
        account_id: str(row.get("type") or "bank_savings")
        for account_id, row in account_map.items()
    }
    category_account_types = {
        "stocks": {"demat"}, "foreign_stocks": {"demat"},
        "mutual_funds": {"mutual_fund"}, "gold": {"gold"},
        "ppf": {"ppf"}, "nps": {"nps"}, "fixed_deposit": {"fixed_deposit"},
    }
    inv_rows = []
    txn_rows = []
    for inv in data:
        if not isinstance(inv, dict) or not inv.get("id") or not inv.get("category"):
            return jsonify({"error": "Each investment requires id and category"}), 400
        try:
            units = float(inv.get("units") or 0)
            buy_price = float(inv.get("buyPrice") or 0)
            current_price = float(inv.get("currentPrice") or 0)
        except (TypeError, ValueError):
            return jsonify({"error": "Investment amounts must be numeric"}), 400
        if min(units, buy_price, current_price) < 0:
            return jsonify({"error": "Investment amounts cannot be negative"}), 400
        entry_mode = str(inv.get("entryMode") or "connected").lower()
        if entry_mode not in {"connected", "prior"}:
            return jsonify({"error": "Invalid investment entry mode"}), 400
        inv["entryMode"] = entry_mode
        container_id = inv.get("containerAccountId")
        if container_id:
            try:
                container_id = int(container_id)
            except (TypeError, ValueError):
                return jsonify({"error": "Invalid investment account"}), 400
            allowed_types = category_account_types.get(str(inv.get("category")), set())
            if account_types.get(container_id) not in allowed_types:
                return jsonify({"error": "Holding is linked to an incompatible investment account"}), 400
            inv["containerAccountId"] = container_id
        inv_rows.append({c: inv.get(c) for c in INVESTMENT_COLS})
        for txn in inv.get("transactions", []):
            action = str(txn.get("action") or "").upper()
            try:
                txn_units = float(txn.get("units") or 0)
                txn_price = float(txn.get("price") or 0)
                txn_charges = float(txn.get("charges") or 0)
            except (TypeError, ValueError):
                return jsonify({"error": "Transaction units, price, and charges must be numeric"}), 400
            valid_actions = {
                "BUY", "SELL", "DEPOSIT", "INTEREST", "WITHDRAWAL", "ADJUSTMENT",
            }
            source = str(txn.get("source") or "connected").lower()
            if source not in {"connected", "opening", "recurring"}:
                return jsonify({"error": "Invalid investment transaction source"}), 400
            if (action not in valid_actions or txn_units <= 0 or txn_charges < 0
                    or (txn_price < 0 and action != "ADJUSTMENT")):
                return jsonify({"error": "Invalid investment transaction"}), 400
            transaction_date = str(txn.get("date") or "")[:10]
            explicit_settlement_date = str(txn.get("settlementDate") or "")[:10]
            settlement_date = explicit_settlement_date or transaction_date
            try:
                datetime.strptime(transaction_date, "%Y-%m-%d")
                datetime.strptime(settlement_date, "%Y-%m-%d")
            except ValueError:
                return jsonify({"error": "Transaction and settlement dates must use YYYY-MM-DD"}), 400
            if settlement_date < transaction_date:
                return jsonify({"error": "Settlement date cannot be before the transaction date"}), 400
            gross_amount = txn_units * txn_price
            if action in {"SELL", "WITHDRAWAL"} and txn_charges > gross_amount:
                return jsonify({"error": "Charges cannot exceed sale or withdrawal proceeds"}), 400
            account_id = None if source == "opening" else txn.get("accountId")
            if account_id:
                try:
                    account_id = int(account_id)
                except (TypeError, ValueError):
                    return jsonify({"error": "Invalid transaction account"}), 400
                if account_id not in account_map:
                    return jsonify({"error": "Transaction account does not exist"}), 400

            transaction_account = account_map.get(account_id, {})
            transaction_account_type = str(transaction_account.get("type") or "")
            transaction_classification = str(transaction_account.get("classification") or "asset").lower()
            category = str(inv.get("category") or "")
            container_id = int(inv.get("containerAccountId") or 0)
            broker_cash_purchase = (
                action == "BUY" and category in {"stocks", "foreign_stocks"}
                and account_id == container_id and account_types.get(container_id) == "demat"
            )
            if action in {"BUY", "DEPOSIT"} and source != "opening":
                if (not account_id or transaction_classification == "liability"
                        or (transaction_classification == "investment" and not broker_cash_purchase)):
                    return jsonify({"error": "Investment funding must come from a bank/cash asset or available Demat broker cash"}), 400
            if action == "WITHDRAWAL" and source != "opening":
                if (not account_id or transaction_classification != "asset"
                        or transaction_account_type in {"credit_card", "loan"}):
                    return jsonify({"error": "Investment withdrawal must credit a bank, cash, or wallet asset account"}), 400

            # New settlement-aware sales must follow the account relationship.
            # Older rows without an explicit settlementDate remain valid.
            if action == "SELL" and txn.get("settlementDate"):
                if category == "mutual_funds":
                    expected_id = int(account_map.get(container_id, {}).get("settlementAccountId") or 0)
                    if not expected_id or account_id != expected_id:
                        return jsonify({"error": "Mutual-fund redemption must credit its linked settlement account"}), 400
                elif category in {"stocks", "foreign_stocks"} and account_id != container_id:
                    return jsonify({"error": "Stock-sale proceeds must first credit the linked Demat account"}), 400
            txn_rows.append({
                "investmentId": inv["id"],
                "date": transaction_date,
                "action": action,
                "units": txn_units,
                "price": txn_price,
                "accountId": account_id,
                "source": source,
                "settlementDate": explicit_settlement_date,
                "charges": txn_charges,
            })

    _write_sheets([
        ("Investments", INVESTMENT_COLS, inv_rows),
        ("Transactions", TRANSACTION_COLS, txn_rows),
    ])
    return jsonify({"ok": True, "count": len(inv_rows)})


# ═══════════════════════════════════════════════════════════════
#  API: SAVINGS HISTORY
# ═══════════════════════════════════════════════════════════════

# ── Helper: compute derived savings columns from other sheets ──
_MONTH_ABBREVS = ["Jan","Feb","Mar","Apr","May","Jun",
                  "Jul","Aug","Sep","Oct","Nov","Dec"]

def _parse_month_label(label):
    """'Apr 2026' → (2026, 4) or None."""
    parts = (label or "").split()
    if len(parts) != 2:
        return None
    try:
        mi = _MONTH_ABBREVS.index(parts[0]) + 1
        yr = int(parts[1])
        return yr, mi
    except (ValueError, IndexError):
        return None


def _compute_derived_savings(rows):
    """Build monthly summaries from dated income, expense, and investment events."""
    income_map = {}
    income_accounts = {}
    for item in _read_sheet("IncomeTransactions", INCOME_COLS):
        d = item.get("date") or ""
        key = d[:7]
        income_map[key] = income_map.get(key, 0) + (float(item["amount"]) if item["amount"] else 0)
        if item.get("accountId"):
            income_accounts.setdefault(key, set()).add(int(item["accountId"]))

    # Expenses by YYYY-MM
    exp_map = {}
    for e in _read_sheet("Expenses", EXPENSE_COLS):
        d = e.get("date") or ""
        key = d[:7]  # "YYYY-MM"
        exp_map[key] = exp_map.get(key, 0) + (float(e["amount"]) if e["amount"] else 0)

    # Investment outflows (BUY transactions) by YYYY-MM
    inv_map = {}
    for t in _read_sheet("Transactions", TRANSACTION_COLS):
        if ((t.get("action") or "").upper() in {"BUY", "DEPOSIT"}
                and str(t.get("source") or "connected").lower() != "opening"):
            d = t.get("date") or ""
            key = d[:7]
            units = float(t["units"]) if t["units"] else 0
            price = float(t["price"]) if t["price"] else 0
            inv_map[key] = inv_map.get(key, 0) + units * price

    # EF contributions by YYYY-MM
    ef_map = {}
    for c in _read_sheet("EFContributions", EMERGENCY_CONTRIB_COLS):
        d = c.get("date") or ""
        key = d[:7]
        ef_map[key] = ef_map.get(key, 0) + (float(c["amount"]) if c["amount"] else 0)

    existing_months = set()
    for r in rows:
        parsed = _parse_month_label(r.get("month"))
        if not parsed:
            continue
        yr, mi = parsed
        ym = f"{yr}-{mi:02d}"
        existing_months.add(ym)
        income  = income_map.get(ym, 0)
        exp     = exp_map.get(ym, 0)
        inv     = inv_map.get(ym, 0)
        ef      = ef_map.get(ym, 0)
        r["income"]    = income
        account_ids = income_accounts.get(ym, set())
        r["accountId"] = next(iter(account_ids)) if len(account_ids) == 1 else None
        r["expenses"]  = exp
        r["invested"]  = inv
        r["emergency"] = ef
        r["net_saved"] = income - exp - inv - ef
    activity_months = set(income_map) | set(exp_map) | set(inv_map) | set(ef_map)
    for ym in sorted(activity_months - existing_months):
        try:
            year, month = (int(value) for value in ym.split("-"))
        except (TypeError, ValueError):
            continue
        income = income_map.get(ym, 0)
        exp = exp_map.get(ym, 0)
        inv = inv_map.get(ym, 0)
        ef = ef_map.get(ym, 0)
        account_ids = income_accounts.get(ym, set())
        rows.append({
            "month": f"{_MONTH_ABBREVS[month - 1]} {year}",
            "income": income,
            "expenses": exp,
            "invested": inv,
            "emergency": ef,
            "net_saved": income - exp - inv - ef,
            "accountId": next(iter(account_ids)) if len(account_ids) == 1 else None,
        })
    rows.sort(key=lambda row: _parse_month_label(row.get("month")) or (0, 0))
    return rows


@app.route("/api/savings-history", methods=["GET"])
def get_savings_history():
    rows = _read_sheet("SavingsHistory", SAVINGS_HIST_COLS)
    rows = _compute_derived_savings(rows)
    return jsonify(rows)


@app.route("/api/savings-history", methods=["POST"])
def save_savings_history():
    data = request.get_json(force=True)
    if not isinstance(data, list):
        return jsonify({"error": "Expected array"}), 400
    # Recompute derived columns before writing
    data = _compute_derived_savings(data)
    _write_sheet("SavingsHistory", SAVINGS_HIST_COLS, data)
    return jsonify({"ok": True, "count": len(data)})


@app.route("/api/savings-history/<month_label>", methods=["PATCH"])
def patch_savings_month(month_label):
    """Update or create income for a single month, e.g. 'Apr 2026'."""
    body = request.get_json(force=True)
    income = body.get("income")
    account_id = body.get("accountId")
    if income is None:
        return jsonify({"error": "income required"}), 400
    rows = _read_sheet("SavingsHistory", SAVINGS_HIST_COLS)
    found = False
    for r in rows:
        if r["month"] == month_label:
            r["income"] = income
            if account_id is not None:
                r["accountId"] = account_id
            found = True
            break
    if not found:
        rows.append({"month": month_label, "income": income, "accountId": account_id})
    rows = _compute_derived_savings(rows)
    _write_sheet("SavingsHistory", SAVINGS_HIST_COLS, rows)
    return jsonify({"ok": True, "month": month_label, "income": income})


# ═══════════════════════════════════════════════════════════════
#  API: SAVINGS GOALS
# ═══════════════════════════════════════════════════════════════

@app.route("/api/savings-goals", methods=["GET"])
def get_savings_goals():
    rows = _read_sheet("SavingsGoals", SAVINGS_GOAL_COLS)
    for r in rows:
        r["id"] = int(r["id"]) if r["id"] else 0
        r["target"] = float(r["target"]) if r["target"] else 0
        r["current"] = float(r["current"]) if r["current"] else 0
    return jsonify(rows)


@app.route("/api/savings-goals", methods=["POST"])
def save_savings_goals():
    data = request.get_json(force=True)
    if not isinstance(data, list):
        return jsonify({"error": "Expected array"}), 400
    _write_sheet("SavingsGoals", SAVINGS_GOAL_COLS, data)
    return jsonify({"ok": True, "count": len(data)})


# ═══════════════════════════════════════════════════════════════
#  API: EMERGENCY FUND
# ═══════════════════════════════════════════════════════════════

@app.route("/api/emergency-fund", methods=["GET"])
def get_emergency_fund():
    rows = _read_sheet("EmergencyFund", EMERGENCY_COLS)
    target = float(rows[0]["target"]) if rows and rows[0]["target"] else 0

    contribs = _read_sheet("EFContributions", EMERGENCY_CONTRIB_COLS)
    contrib_list = []
    for c in contribs:
        contrib_list.append({
            "id":     int(c["id"]) if c["id"] else 0,
            "date":   str(c["date"]) if c["date"] else "",
            "amount": float(c["amount"]) if c["amount"] else 0,
            "note":   str(c["note"]) if c["note"] else "",
        })
    current = sum(c["amount"] for c in contrib_list)
    return jsonify({"target": target, "current": current, "contributions": contrib_list})


@app.route("/api/emergency-fund", methods=["POST"])
def save_emergency_fund():
    data = request.get_json(force=True)
    if not isinstance(data, dict):
        return jsonify({"error": "Expected object"}), 400
    try:
        target = float(data.get("target") or 0)
    except (TypeError, ValueError):
        return jsonify({"error": "Emergency target must be numeric"}), 400
    if not math.isfinite(target):
        return jsonify({"error": "Emergency target must be a finite number"}), 400
    if target < 0:
        return jsonify({"error": "Emergency target cannot be negative"}), 400
    contribs = data.get("contributions", [])
    if not isinstance(contribs, list):
        return jsonify({"error": "Emergency contributions must be an array"}), 400
    rows = []
    for c in contribs:
        if not isinstance(c, dict):
            return jsonify({"error": "Invalid legacy emergency contribution"}), 400
        try:
            contribution_id = int(c.get("id", 0))
            amount = float(c.get("amount") or 0)
        except (TypeError, ValueError):
            return jsonify({"error": "Legacy emergency contribution must be numeric"}), 400
        if not math.isfinite(amount):
            return jsonify({"error": "Legacy emergency contribution must be a finite number"}), 400
        rows.append({
            "id":     contribution_id,
            "date":   c.get("date", ""),
            "amount": amount,
            "note":   str(c.get("note") or "")[:160],
        })
    _write_sheets([
        ("EmergencyFund", EMERGENCY_COLS, [{"target": target}]),
        ("EFContributions", EMERGENCY_CONTRIB_COLS, rows),
    ])
    return jsonify({"ok": True})


@app.route("/api/emergency-allocations", methods=["GET"])
def get_emergency_allocations():
    rows = _read_sheet("EmergencyAllocations", EMERGENCY_ALLOCATION_COLS)
    for row in rows:
        row["id"] = int(row["id"]) if row.get("id") else 0
        row["sourceId"] = int(row["sourceId"]) if row.get("sourceId") else 0
        row["amount"] = float(row["amount"] or 0)
        row["sourceType"] = str(row.get("sourceType") or "")
        row["allocationMode"] = str(row.get("allocationMode") or "amount")
        row["liquidity"] = str(row.get("liquidity") or "redeemable")
        row["note"] = str(row.get("note") or "")
        row["updatedAt"] = str(row.get("updatedAt") or "")
    return jsonify(rows)


@app.route("/api/emergency-allocations", methods=["POST"])
def save_emergency_allocations():
    data = request.get_json(force=True)
    if not isinstance(data, list):
        return jsonify({"error": "Expected array"}), 400

    accounts = {
        int(row["id"]): row for row in _read_sheet("Accounts", ACCOUNT_COLS)
        if row.get("id")
    }
    investments = {
        int(row["id"]): row for row in _read_sheet("Investments", INVESTMENT_COLS)
        if row.get("id")
    }
    investment_types = {"demat", "mutual_fund", "gold", "ppf", "nps", "fixed_deposit"}
    cleaned = []
    seen_ids = set()
    seen_sources = set()
    now = datetime.now().isoformat()

    for row in data:
        if not isinstance(row, dict):
            return jsonify({"error": "Invalid emergency allocation"}), 400
        try:
            row_id = int(row.get("id"))
            source_id = int(row.get("sourceId"))
            amount = float(row.get("amount") or 0)
        except (TypeError, ValueError):
            return jsonify({"error": "Allocation requires numeric IDs and amount"}), 400
        source_type = str(row.get("sourceType") or "").lower()
        allocation_mode = str(row.get("allocationMode") or "amount").lower()
        liquidity = str(row.get("liquidity") or "").lower()
        source_key = (source_type, source_id)

        if not math.isfinite(amount):
            return jsonify({"error": "Allocation amount must be a finite number"}), 400
        if row_id <= 0 or row_id in seen_ids:
            return jsonify({"error": "Emergency allocations require unique positive IDs"}), 400
        if source_type not in {"account", "investment"}:
            return jsonify({"error": "Allocation source must be an account or investment"}), 400
        if source_key in seen_sources:
            return jsonify({"error": "An account or holding can be allocated only once"}), 400
        if allocation_mode not in {"full", "amount"}:
            return jsonify({"error": "Allocation mode must be full or amount"}), 400
        if liquidity not in {"immediate", "redeemable", "locked"}:
            return jsonify({"error": "Invalid emergency-fund liquidity"}), 400
        if allocation_mode == "amount" and amount <= 0:
            return jsonify({"error": "A fixed allocation must be greater than zero"}), 400
        if amount < 0:
            return jsonify({"error": "Allocation amount cannot be negative"}), 400

        if source_type == "account":
            account = accounts.get(source_id)
            account_type = str((account or {}).get("type") or "").lower()
            classification = str((account or {}).get("classification") or (
                "liability" if account_type in {"credit_card", "loan"}
                else "investment" if account_type in investment_types
                else "asset"
            )).lower()
            if not account or classification != "asset":
                return jsonify({"error": "Only bank, cash, or wallet assets can fund an emergency reserve"}), 400
        elif source_id not in investments:
            return jsonify({"error": "Emergency allocation references a missing investment"}), 400

        seen_ids.add(row_id)
        seen_sources.add(source_key)
        cleaned.append({
            "id": row_id,
            "sourceType": source_type,
            "sourceId": source_id,
            "allocationMode": allocation_mode,
            "amount": 0 if allocation_mode == "full" else amount,
            "liquidity": liquidity,
            "note": str(row.get("note") or "").strip()[:160],
            "updatedAt": str(row.get("updatedAt") or now)[:32],
        })

    _write_sheet("EmergencyAllocations", EMERGENCY_ALLOCATION_COLS, cleaned)
    return jsonify({"ok": True, "count": len(cleaned)})


# ═══════════════════════════════════════════════════════════════
#  API: FINANCIAL PLANNING
# ═══════════════════════════════════════════════════════════════

def _planning_response(sheet_name, columns):
    return jsonify(_read_sheet(sheet_name, columns))


def _save_planning_rows(sheet_name, columns):
    data = request.get_json(force=True)
    if not isinstance(data, list):
        return jsonify({"error": "Expected array"}), 400
    cleaned = [{column: row.get(column) for column in columns} for row in data if isinstance(row, dict)]
    _write_sheet(sheet_name, columns, cleaned)
    return jsonify({"ok": True, "count": len(cleaned)})


@app.route("/api/budgets", methods=["GET"])
def get_budgets():
    return _planning_response("Budgets", BUDGET_COLS)


@app.route("/api/budgets", methods=["POST"])
def save_budgets():
    return _save_planning_rows("Budgets", BUDGET_COLS)


@app.route("/api/recurring-bills", methods=["GET"])
def get_recurring_bills():
    return _planning_response("RecurringBills", RECURRING_BILL_COLS)


@app.route("/api/recurring-bills", methods=["POST"])
def save_recurring_bills():
    return _save_planning_rows("RecurringBills", RECURRING_BILL_COLS)


@app.route("/api/net-worth", methods=["GET"])
def get_net_worth():
    return _planning_response("NetWorth", NET_WORTH_COLS)


@app.route("/api/net-worth", methods=["POST"])
def save_net_worth():
    return _save_planning_rows("NetWorth", NET_WORTH_COLS)


@app.route("/api/net-worth-auto", methods=["GET"])
def get_automatic_net_worth():
    return _planning_response("NetWorthAuto", NET_WORTH_AUTO_COLS)


@app.route("/api/net-worth-auto", methods=["POST"])
def save_automatic_net_worth():
    return _save_planning_rows("NetWorthAuto", NET_WORTH_AUTO_COLS)


@app.route("/api/cash-flow", methods=["GET"])
def get_cash_flow():
    return _planning_response("CashFlow", CASH_FLOW_COLS)


@app.route("/api/cash-flow", methods=["POST"])
def save_cash_flow():
    return _save_planning_rows("CashFlow", CASH_FLOW_COLS)


@app.route("/api/accounts", methods=["GET"])
def get_accounts():
    rows = _read_sheet("Accounts", ACCOUNT_COLS)
    for row in rows:
        row["id"] = int(row["id"]) if row.get("id") else 0
        row["openingBalance"] = float(row["openingBalance"] or 0)
        row["statementBalance"] = float(row["statementBalance"] or 0)
        row["creditLimit"] = float(row["creditLimit"] or 0)
        row["settlementAccountId"] = (
            int(row["settlementAccountId"]) if row.get("settlementAccountId") else None
        )
        row["type"] = str(row.get("type") or "bank_savings")
        row["classification"] = str(row.get("classification") or (
            "liability" if row["type"] in {"credit_card", "loan"}
            else "investment" if row["type"] in {
                "demat", "mutual_fund", "gold", "ppf", "nps", "fixed_deposit"
            } else "asset"
        ))
        row["currency"] = str(row.get("currency") or "INR")
    return jsonify(rows)


@app.route("/api/accounts", methods=["POST"])
def save_accounts():
    data = request.get_json(force=True)
    if not isinstance(data, list):
        return jsonify({"error": "Expected array"}), 400
    cleaned = []
    valid_purposes = {"salary", "investment", "spending", "savings", "other"}
    valid_types = {
        "bank_savings", "bank_current", "cash", "credit_card", "wallet",
        "store", "demat", "mutual_fund", "gold", "ppf", "nps", "fixed_deposit",
        "loan", "other",
    }
    investment_types = {"demat", "mutual_fund", "gold", "ppf", "nps", "fixed_deposit"}
    for row in data:
        if not isinstance(row, dict) or not row.get("id") or not str(row.get("name") or "").strip():
            return jsonify({"error": "Each account requires id and name"}), 400
        purpose = str(row.get("purpose") or "other").lower()
        if purpose not in valid_purposes:
            return jsonify({"error": "Invalid account purpose"}), 400
        account_type = str(row.get("type") or "bank_savings").lower()
        if account_type not in valid_types:
            return jsonify({"error": "Invalid account type"}), 400
        classification = (
            "liability" if account_type in {"credit_card", "loan"}
            else "investment" if account_type in investment_types
            else "asset"
        )
        try:
            opening = float(row.get("openingBalance") or 0)
            statement = float(row.get("statementBalance") or 0)
            credit_limit = float(row.get("creditLimit") or 0)
        except (TypeError, ValueError):
            return jsonify({"error": "Account balances must be numeric"}), 400
        opening_date = str(row.get("openingDate") or "").strip()
        if opening_date:
            try:
                datetime.strptime(opening_date, "%Y-%m-%d")
            except ValueError:
                return jsonify({"error": "Tracking start date must use YYYY-MM-DD"}), 400
        cleaned.append({
            **{column: row.get(column) for column in ACCOUNT_COLS},
            "name": str(row["name"]).strip(),
            "bank": str(row.get("bank") or "").strip(),
            "type": account_type,
            "classification": classification,
            "purpose": purpose,
            "currency": str(row.get("currency") or "INR").upper(),
            "openingDate": opening_date,
            "openingBalance": opening,
            "statementBalance": statement,
            "creditLimit": credit_limit,
            "settlementAccountId": row.get("settlementAccountId") or None,
        })
    cleaned_by_id = {int(row["id"]): row for row in cleaned}
    settlement_types = {"demat", "mutual_fund"}
    for row in cleaned:
        raw_settlement_id = row.get("settlementAccountId")
        if row["type"] not in settlement_types:
            row["settlementAccountId"] = None
            continue
        if not raw_settlement_id:
            row["settlementAccountId"] = None
            continue
        try:
            settlement_id = int(raw_settlement_id)
        except (TypeError, ValueError):
            return jsonify({"error": "Invalid linked settlement account"}), 400
        settlement_account = cleaned_by_id.get(settlement_id)
        if (not settlement_account or settlement_id == int(row["id"])
                or settlement_account["classification"] != "asset"):
            return jsonify({"error": "Settlement account must be a bank, cash, or wallet asset account"}), 400
        row["settlementAccountId"] = settlement_id
    try:
        _write_sheet("Accounts", ACCOUNT_COLS, cleaned)
        persisted = _read_sheet("Accounts", ACCOUNT_COLS)
    except PermissionError:
        return jsonify({
            "error": "data.xlsx is locked. Close it in Excel/OneDrive and try again."
        }), 423
    except OSError as exc:
        return jsonify({"error": f"Could not save data.xlsx: {exc}"}), 500

    expected_ids = {str(row["id"]) for row in cleaned}
    persisted_ids = {str(row.get("id")) for row in persisted}
    if len(persisted) != len(cleaned) or persisted_ids != expected_ids:
        return jsonify({
            "error": "Account save could not be verified in data.xlsx. Please try again."
        }), 500
    return jsonify({"ok": True, "count": len(persisted), "verified": True})


@app.route("/api/transfers", methods=["GET"])
def get_transfers():
    rows = _read_sheet("Transfers", TRANSFER_COLS)
    for row in rows:
        row["id"] = int(row["id"]) if row.get("id") else 0
        row["fromAccountId"] = int(row["fromAccountId"]) if row.get("fromAccountId") else None
        row["toAccountId"] = int(row["toAccountId"]) if row.get("toAccountId") else None
        row["amount"] = float(row["amount"] or 0)
    return jsonify(rows)


@app.route("/api/transfers", methods=["POST"])
def save_transfers():
    data = request.get_json(force=True)
    if not isinstance(data, list):
        return jsonify({"error": "Expected array"}), 400
    account_ids = {
        int(row["id"]) for row in _read_sheet("Accounts", ACCOUNT_COLS) if row.get("id")
    }
    cleaned = []
    for row in data:
        try:
            from_id = int(row.get("fromAccountId"))
            to_id = int(row.get("toAccountId"))
            amount = float(row.get("amount") or 0)
        except (TypeError, ValueError, AttributeError):
            return jsonify({"error": "Invalid transfer"}), 400
        if (from_id == to_id or from_id not in account_ids or to_id not in account_ids
                or amount <= 0 or not row.get("date")):
            return jsonify({"error": "Transfer requires different accounts, date, and positive amount"}), 400
        cleaned.append({
            "id": row.get("id"),
            "date": row.get("date"),
            "fromAccountId": from_id,
            "toAccountId": to_id,
            "amount": amount,
            "note": str(row.get("note") or "").strip(),
        })
    _write_sheet("Transfers", TRANSFER_COLS, cleaned)
    return jsonify({"ok": True, "count": len(cleaned)})


# ═══════════════════════════════════════════════════════════════
#  API: ACCOUNT RECONCILIATION ADJUSTMENTS
# =============================================================================

@app.route("/api/reconciliation-adjustments", methods=["GET"])
def get_reconciliation_adjustments():
    rows = _read_sheet("ReconciliationAdjustments", RECONCILIATION_ADJUSTMENT_COLS)
    for row in rows:
        row["accountId"] = int(row["accountId"]) if row.get("accountId") else None
        row["amount"] = float(row["amount"] or 0)
    return jsonify(rows)


@app.route("/api/reconciliation-adjustments", methods=["POST"])
def save_reconciliation_adjustments():
    data = request.get_json(force=True)
    if not isinstance(data, list):
        return jsonify({"error": "Expected array"}), 400
    account_ids = {
        int(row["id"]) for row in _read_sheet("Accounts", ACCOUNT_COLS) if row.get("id")
    }
    cleaned = []
    for row in data:
        try:
            account_id = int(row.get("accountId"))
            amount = float(row.get("amount"))
            datetime.strptime(str(row.get("date")), "%Y-%m-%d")
        except (TypeError, ValueError, AttributeError):
            return jsonify({"error": "Invalid reconciliation adjustment"}), 400
        reason = str(row.get("reason") or "").strip()
        if account_id not in account_ids or amount == 0 or not reason:
            return jsonify({
                "error": "Adjustment requires a valid account, non-zero amount, date, and reason"
            }), 400
        cleaned.append({
            "id": str(row.get("id") or ""),
            "accountId": account_id,
            "date": row.get("date"),
            "amount": amount,
            "reason": reason,
            "createdAt": str(row.get("createdAt") or datetime.now().isoformat(timespec="seconds")),
        })
    _write_sheet("ReconciliationAdjustments", RECONCILIATION_ADJUSTMENT_COLS, cleaned)
    return jsonify({"ok": True, "count": len(cleaned)})


#  API: RECURRING INVESTMENT RULES / CATCH-UP OCCURRENCES
# =============================================================================

def _scheduled_dates(rule, through_date):
    """Return every due date for a rule through the supplied date."""
    try:
        start = _to_date(rule.get("startDate"))
        end = _to_date(rule.get("endDate")) if rule.get("endDate") else through_date
        day = max(1, min(31, int(rule.get("day") or start.day)))
    except (TypeError, ValueError, AttributeError):
        return []
    if not isinstance(start, date):
        return []
    end = min(end, through_date) if isinstance(end, date) else through_date
    step = {"monthly": 1, "quarterly": 3, "yearly": 12}.get(
        str(rule.get("frequency") or "monthly").lower()
    )
    if not step:
        return []
    year, month = start.year, start.month
    dates = []
    while date(year, month, 1) <= end:
        scheduled = date(year, month, min(day, calendar.monthrange(year, month)[1]))
        if start <= scheduled <= end:
            dates.append(scheduled)
        month += step
        year += (month - 1) // 12
        month = (month - 1) % 12 + 1
    return dates


@app.route("/api/recurring-rules", methods=["GET"])
def get_recurring_rules():
    rows = _read_sheet("RecurringRules", RECURRING_RULE_COLS)
    for row in rows:
        row["id"] = int(row["id"]) if row.get("id") else 0
        row["day"] = int(row["day"] or 1)
        row["amount"] = float(row["amount"] or 0)
        row["fromAccountId"] = int(row["fromAccountId"]) if row.get("fromAccountId") else None
        row["investmentId"] = int(row["investmentId"]) if row.get("investmentId") else None
    return jsonify(rows)


@app.route("/api/recurring-rules", methods=["POST"])
def save_recurring_rules():
    data = request.get_json(force=True)
    if not isinstance(data, list):
        return jsonify({"error": "Expected array"}), 400
    accounts = {
        int(row["id"]): row for row in _read_sheet("Accounts", ACCOUNT_COLS) if row.get("id")
    }
    investment_ids = {int(row["id"]) for row in _read_sheet("Investments", INVESTMENT_COLS) if row.get("id")}
    cleaned = []
    for row in data:
        try:
            rule_id = int(row.get("id")); day = int(row.get("day"))
            amount = float(row.get("amount")); from_id = int(row.get("fromAccountId"))
            investment_id = int(row.get("investmentId"))
            datetime.strptime(str(row.get("startDate")), "%Y-%m-%d")
        except (TypeError, ValueError, AttributeError):
            return jsonify({"error": "Invalid recurring rule"}), 400
        frequency = str(row.get("frequency") or "monthly").lower()
        if (not str(row.get("name") or "").strip() or day < 1 or day > 31
                or amount <= 0 or from_id not in accounts
                or investment_id not in investment_ids
                or frequency not in {"monthly", "quarterly", "yearly"}):
            return jsonify({"error": "Recurring rule has invalid fields"}), 400
        funding_account = accounts[from_id]
        if str(funding_account.get("classification") or "asset").lower() != "asset":
            return jsonify({"error": "Recurring investments must be funded from a bank, cash, or wallet asset account"}), 400
        cleaned.append({
            "id": rule_id, "name": str(row["name"]).strip(), "type": "sip",
            "frequency": frequency, "day": day, "amount": amount,
            "fromAccountId": from_id, "investmentId": investment_id,
            "startDate": row.get("startDate"), "endDate": row.get("endDate") or "",
            "active": row.get("active") is not False,
        })
    _write_sheet("RecurringRules", RECURRING_RULE_COLS, cleaned)
    return jsonify({"ok": True, "count": len(cleaned)})


@app.route("/api/recurring-occurrences", methods=["GET"])
def get_recurring_occurrences():
    rows = _read_sheet("RecurringOccurrences", RECURRING_OCCURRENCE_COLS)
    for row in rows:
        row["ruleId"] = int(row["ruleId"]) if row.get("ruleId") else 0
        for column in ("actualAmount", "units", "price"):
            row[column] = float(row[column] or 0)
    return jsonify(rows)


@app.route("/api/recurring-occurrences/generate", methods=["POST"])
def generate_recurring_occurrences():
    rules = _read_sheet("RecurringRules", RECURRING_RULE_COLS)
    rows = _read_sheet("RecurringOccurrences", RECURRING_OCCURRENCE_COLS)
    existing = {str(row.get("id")) for row in rows}; created = 0
    for rule in rules:
        if rule.get("active") is False or str(rule.get("active")).lower() == "false":
            continue
        rule_id = int(rule.get("id") or 0)
        for scheduled in _scheduled_dates(rule, date.today()):
            occurrence_id = f"{rule_id}:{scheduled.isoformat()}"
            if occurrence_id in existing:
                continue
            rows.append({
                "id": occurrence_id, "ruleId": rule_id,
                "scheduledDate": scheduled.isoformat(), "status": "pending",
                "actualDate": "", "actualAmount": 0, "units": 0, "price": 0, "note": "",
            })
            existing.add(occurrence_id); created += 1
    if created:
        _write_sheet("RecurringOccurrences", RECURRING_OCCURRENCE_COLS, rows)
    pending = sum(1 for row in rows if str(row.get("status") or "").lower() == "pending")
    return jsonify({"ok": True, "created": created, "pending": pending})


@app.route("/api/recurring-occurrences/<path:occurrence_id>/action", methods=["POST"])
def recurring_occurrence_action(occurrence_id):
    body = request.get_json(force=True); action = str(body.get("action") or "").lower()
    occurrences = _read_sheet("RecurringOccurrences", RECURRING_OCCURRENCE_COLS)
    occurrence = next((row for row in occurrences if str(row.get("id")) == occurrence_id), None)
    if not occurrence or str(occurrence.get("status") or "").lower() != "pending":
        return jsonify({"error": "Pending occurrence not found"}), 404
    if action == "skip":
        occurrence["status"] = "skipped"; occurrence["note"] = str(body.get("note") or "").strip()
        _write_sheet("RecurringOccurrences", RECURRING_OCCURRENCE_COLS, occurrences)
        return jsonify({"ok": True, "status": "skipped"})
    if action != "confirm":
        return jsonify({"error": "Action must be confirm or skip"}), 400
    rules = _read_sheet("RecurringRules", RECURRING_RULE_COLS)
    rule = next((row for row in rules if int(row.get("id") or 0) == int(occurrence.get("ruleId") or 0)), None)
    if not rule:
        return jsonify({"error": "Recurring rule not found"}), 404
    try:
        actual_date = str(body.get("actualDate") or occurrence.get("scheduledDate"))
        datetime.strptime(actual_date, "%Y-%m-%d")
        amount = float(body.get("actualAmount") or rule.get("amount") or 0)
        price = float(body.get("price") or 0)
        units = float(body.get("units") or (amount / price if price > 0 else 1))
    except (TypeError, ValueError, ZeroDivisionError):
        return jsonify({"error": "Invalid confirmation values"}), 400
    if amount <= 0 or units <= 0 or price <= 0:
        return jsonify({"error": "Amount, units, and price must be positive"}), 400
    investments = _read_sheet("Investments", INVESTMENT_COLS)
    transactions = _read_sheet("Transactions", TRANSACTION_COLS)
    investment_id = int(rule.get("investmentId") or 0)
    investment = next((row for row in investments if int(row.get("id") or 0) == investment_id), None)
    if not investment:
        return jsonify({"error": "Linked investment not found"}), 404
    category = str(investment.get("category") or "")
    action_name = "DEPOSIT" if category in {"ppf", "fixed_deposit"} else "BUY"
    if action_name == "BUY" and abs((units * price) - amount) > max(1.0, amount * 0.001):
        return jsonify({"error": "Amount must match units multiplied by price"}), 400
    txn_units = 1 if action_name == "DEPOSIT" else units
    txn_price = amount if action_name == "DEPOSIT" else price
    transactions.append({
        "investmentId": investment_id, "date": actual_date, "action": action_name,
        "units": txn_units, "price": txn_price, "accountId": int(rule.get("fromAccountId")),
        "source": "recurring",
    })
    if action_name == "BUY":
        old_units = float(investment.get("units") or 0)
        old_cost = old_units * float(investment.get("buyPrice") or 0)
        investment["units"] = old_units + units
        investment["buyPrice"] = (old_cost + amount) / investment["units"]
    occurrence.update({
        "status": "confirmed", "actualDate": actual_date, "actualAmount": amount,
        "units": txn_units, "price": txn_price, "note": str(body.get("note") or "").strip(),
    })
    _write_sheets([
        ("Investments", INVESTMENT_COLS, investments),
        ("Transactions", TRANSACTION_COLS, transactions),
        ("RecurringOccurrences", RECURRING_OCCURRENCE_COLS, occurrences),
    ])
    return jsonify({"ok": True, "status": "confirmed"})


#  MUTUAL-FUND CATALOGUE & NAV CACHE
# ═══════════════════════════════════════════════════════════════

MFAPI_BASE_URL = "https://api.mfapi.in"


def _market_cache_connection():
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(MARKET_CACHE_FILE), timeout=15)
    conn.row_factory = sqlite3.Row
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS mf_schemes (
            scheme_code TEXT PRIMARY KEY,
            scheme_name TEXT NOT NULL,
            fund_house TEXT,
            scheme_type TEXT,
            scheme_category TEXT,
            isin_growth TEXT,
            isin_div_reinvestment TEXT,
            updated_at TEXT NOT NULL
        );
        CREATE INDEX IF NOT EXISTS idx_mf_schemes_name
            ON mf_schemes(scheme_name COLLATE NOCASE);
        CREATE TABLE IF NOT EXISTS mf_nav_cache (
            scheme_code TEXT NOT NULL,
            nav_date TEXT NOT NULL,
            nav REAL NOT NULL,
            fetched_at TEXT NOT NULL,
            PRIMARY KEY (scheme_code, nav_date)
        );
        CREATE TABLE IF NOT EXISTS mf_cache_meta (
            key TEXT PRIMARY KEY,
            value TEXT
        );
    """)
    return conn


def _normalise_mf_scheme(item):
    if not isinstance(item, dict):
        return None
    code = item.get("schemeCode") or item.get("scheme_code")
    name = item.get("schemeName") or item.get("scheme_name")
    if code is None or not name:
        return None
    return {
        "schemeCode": str(code).strip(),
        "schemeName": str(name).strip(),
        "fundHouse": item.get("fund_house") or item.get("fundHouse"),
        "schemeType": item.get("scheme_type") or item.get("schemeType"),
        "schemeCategory": item.get("scheme_category") or item.get("schemeCategory"),
        "isinGrowth": item.get("isin_growth") or item.get("isinGrowth"),
        "isinDivReinvestment": (
            item.get("isin_div_reinvestment") or item.get("isinDivReinvestment")
        ),
    }


def _upsert_mf_schemes(conn, items):
    now = datetime.now().isoformat(timespec="seconds")
    rows = [row for row in (_normalise_mf_scheme(item) for item in items) if row]
    conn.executemany("""
        INSERT INTO mf_schemes (
            scheme_code, scheme_name, fund_house, scheme_type, scheme_category,
            isin_growth, isin_div_reinvestment, updated_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(scheme_code) DO UPDATE SET
            scheme_name = excluded.scheme_name,
            fund_house = COALESCE(excluded.fund_house, mf_schemes.fund_house),
            scheme_type = COALESCE(excluded.scheme_type, mf_schemes.scheme_type),
            scheme_category = COALESCE(excluded.scheme_category, mf_schemes.scheme_category),
            isin_growth = COALESCE(excluded.isin_growth, mf_schemes.isin_growth),
            isin_div_reinvestment = COALESCE(
                excluded.isin_div_reinvestment, mf_schemes.isin_div_reinvestment
            ),
            updated_at = excluded.updated_at
    """, [(
        row["schemeCode"], row["schemeName"], row["fundHouse"], row["schemeType"],
        row["schemeCategory"], row["isinGrowth"], row["isinDivReinvestment"], now,
    ) for row in rows])
    return len(rows)


def _mf_catalog_status():
    with _market_cache_lock:
        conn = _market_cache_connection()
        try:
            count = conn.execute("SELECT COUNT(*) FROM mf_schemes").fetchone()[0]
            meta = conn.execute(
                "SELECT value FROM mf_cache_meta WHERE key = 'catalog_refreshed_at'"
            ).fetchone()
            return {"count": int(count), "lastRefreshed": meta[0] if meta else None}
        finally:
            conn.close()


def _refresh_mf_catalog():
    all_items = []
    seen_codes = set()
    limit = 1000
    offset = 0
    for _page in range(100):
        response = requests.get(
            f"{MFAPI_BASE_URL}/mf", params={"limit": limit, "offset": offset}, timeout=25
        )
        response.raise_for_status()
        payload = response.json()
        if isinstance(payload, list):
            items = payload
        elif isinstance(payload, dict):
            items = payload.get("data") or payload.get("schemes") or payload.get("results") or []
        else:
            items = []
        normalised = [row for row in (_normalise_mf_scheme(item) for item in items) if row]
        new_items = [row for row in normalised if row["schemeCode"] not in seen_codes]
        all_items.extend(new_items)
        seen_codes.update(row["schemeCode"] for row in new_items)
        if not items or len(items) < limit or not new_items:
            break
        offset += len(items)
    if not all_items:
        raise RuntimeError("MFapi returned no schemes")

    refreshed_at = datetime.now().isoformat(timespec="seconds")
    with _market_cache_lock:
        conn = _market_cache_connection()
        try:
            _upsert_mf_schemes(conn, all_items)
            conn.execute("""
                INSERT INTO mf_cache_meta(key, value) VALUES('catalog_refreshed_at', ?)
                ON CONFLICT(key) DO UPDATE SET value = excluded.value
            """, (refreshed_at,))
            conn.commit()
            count = conn.execute("SELECT COUNT(*) FROM mf_schemes").fetchone()[0]
        finally:
            conn.close()
    return {"count": int(count), "updated": len(all_items), "lastRefreshed": refreshed_at}


def _search_cached_mf_schemes(query, limit=20):
    terms = [term.casefold() for term in re.split(r"\s+", query.strip()) if term]
    if not terms:
        return []
    clauses = ["LOWER(scheme_name) LIKE ?" for _ in terms]
    values = [f"%{term}%" for term in terms]
    prefix = f"{query.strip().casefold()}%"
    sql = f"""
        SELECT scheme_code, scheme_name, fund_house, scheme_type, scheme_category,
               isin_growth, isin_div_reinvestment
        FROM mf_schemes
        WHERE {' AND '.join(clauses)}
        ORDER BY CASE WHEN LOWER(scheme_name) LIKE ? THEN 0 ELSE 1 END,
                 scheme_name COLLATE NOCASE
        LIMIT ?
    """
    with _market_cache_lock:
        conn = _market_cache_connection()
        try:
            rows = conn.execute(sql, [*values, prefix, limit]).fetchall()
        finally:
            conn.close()
    return [{
        "schemeCode": row["scheme_code"], "schemeName": row["scheme_name"],
        "fundHouse": row["fund_house"], "schemeType": row["scheme_type"],
        "schemeCategory": row["scheme_category"], "isinGrowth": row["isin_growth"],
        "isinDivReinvestment": row["isin_div_reinvestment"],
    } for row in rows]


def _fetch_remote_mf_search(query):
    response = requests.get(
        f"{MFAPI_BASE_URL}/mf/search", params={"q": query}, timeout=12
    )
    response.raise_for_status()
    payload = response.json()
    return payload if isinstance(payload, list) else []


def _cached_mf_nav(scheme_code):
    with _market_cache_lock:
        conn = _market_cache_connection()
        try:
            row = conn.execute("""
                SELECT n.nav, n.nav_date, n.fetched_at, s.scheme_name
                FROM mf_nav_cache n
                LEFT JOIN mf_schemes s ON s.scheme_code = n.scheme_code
                WHERE n.scheme_code = ?
                ORDER BY n.fetched_at DESC LIMIT 1
            """, (str(scheme_code),)).fetchone()
        finally:
            conn.close()
    if not row:
        return None
    return {
        "nav": float(row["nav"]), "date": row["nav_date"],
        "name": row["scheme_name"], "cached": True, "fetchedAt": row["fetched_at"],
    }


def _cache_mf_nav(scheme_code, nav, nav_date, meta=None):
    fetched_at = datetime.now().isoformat(timespec="seconds")
    with _market_cache_lock:
        conn = _market_cache_connection()
        try:
            if meta:
                _upsert_mf_schemes(conn, [{**meta, "schemeCode": scheme_code}])
            conn.execute("""
                INSERT INTO mf_nav_cache(scheme_code, nav_date, nav, fetched_at)
                VALUES (?, ?, ?, ?)
                ON CONFLICT(scheme_code, nav_date) DO UPDATE SET
                    nav = excluded.nav, fetched_at = excluded.fetched_at
            """, (str(scheme_code), str(nav_date), float(nav), fetched_at))
            conn.commit()
        finally:
            conn.close()


@app.route("/api/mutual-funds/catalog/status")
def mf_catalog_status():
    return jsonify(_mf_catalog_status())


@app.route("/api/mutual-funds/catalog/refresh", methods=["POST"])
def refresh_mf_catalog():
    try:
        return jsonify({"ok": True, **_refresh_mf_catalog()})
    except Exception as exc:
        return jsonify({"error": str(exc), **_mf_catalog_status()}), 502


@app.route("/api/mutual-funds/search")
def search_mf_catalog():
    query = str(request.args.get("q") or "").strip()
    if len(query) < 2:
        return jsonify({"items": [], **_mf_catalog_status()})
    try:
        limit = max(5, min(50, int(request.args.get("limit") or 20)))
    except (TypeError, ValueError):
        limit = 20
    items = _search_cached_mf_schemes(query, limit)
    status = _mf_catalog_status()
    source = "cache"
    if not items and status["count"] == 0:
        try:
            remote_items = _fetch_remote_mf_search(query)
            with _market_cache_lock:
                conn = _market_cache_connection()
                try:
                    _upsert_mf_schemes(conn, remote_items)
                    conn.commit()
                finally:
                    conn.close()
            items = _search_cached_mf_schemes(query, limit)
            status = _mf_catalog_status()
            source = "online"
        except Exception:
            source = "offline-empty"
    return jsonify({"items": items, "source": source, **status})


# ═══════════════════════════════════════════════════════════════
#  API: PRICE PROXY  (solves CORS — Python fetches directly)
# ═══════════════════════════════════════════════════════════════

@app.route("/api/price/stock/<ticker>")
def proxy_stock_price(ticker):
    """Proxy Yahoo Finance chart API for a given ticker."""
    url = f"https://query1.finance.yahoo.com/v8/finance/chart/{ticker}?interval=1d&range=1d"
    headers = {"User-Agent": "Mozilla/5.0"}
    try:
        resp = requests.get(url, headers=headers, timeout=15)
        if resp.status_code != 200:
            return jsonify({"error": f"Yahoo returned {resp.status_code}"}), 502
        data = resp.json()
        meta = data.get("chart", {}).get("result", [{}])[0].get("meta", {})
        price = meta.get("regularMarketPrice")
        return jsonify({
            "price": price,
            "name": meta.get("longName") or meta.get("shortName"),
            "currency": meta.get("currency"),
            "high": meta.get("regularMarketDayHigh"),
            "low": meta.get("regularMarketDayLow"),
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 502


@app.route("/api/price/mf/<scheme_code>")
def proxy_mf_price(scheme_code):
    """Return live MFapi NAV, falling back to the last locally cached NAV."""
    if not re.fullmatch(r"\d{3,12}", str(scheme_code)):
        return jsonify({"error": "Invalid mutual-fund scheme code"}), 400
    try:
        resp = requests.get(f"{MFAPI_BASE_URL}/mf/{scheme_code}/latest", timeout=12)
        if resp.status_code == 404:
            resp = requests.get(f"{MFAPI_BASE_URL}/mf/{scheme_code}", timeout=12)
        resp.raise_for_status()
        data = resp.json()
        nav_entry = (data.get("data") or [{}])[0]
        nav = float(nav_entry.get("nav") or 0)
        nav_date = nav_entry.get("date")
        meta = data.get("meta") or {}
        if nav <= 0 or not nav_date:
            raise RuntimeError("MFapi returned no NAV")
        _cache_mf_nav(scheme_code, nav, nav_date, meta)
        return jsonify({
            "nav": nav, "date": nav_date,
            "name": meta.get("scheme_name"), "cached": False,
            "fetchedAt": datetime.now().isoformat(timespec="seconds"),
        })
    except Exception as exc:
        cached = _cached_mf_nav(scheme_code)
        if cached:
            return jsonify({**cached, "offline": True, "warning": str(exc)})
        return jsonify({"error": str(exc)}), 502


# ═══════════════════════════════════════════════════════════════
#  API: EXPORT / DOWNLOAD EXCEL
# ═══════════════════════════════════════════════════════════════

@app.route("/api/export")
def export_excel():
    """Download the data.xlsx file."""
    return send_from_directory(str(DATA_FILE.parent), DATA_FILE.name,
                               as_attachment=True,
                               download_name="FinTrack_data.xlsx")


# ═══════════════════════════════════════════════════════════════
#  API: DOCUMENTS
# ═══════════════════════════════════════════════════════════════

DOCS_DIR = BASE_DIR / "documents"
DEFAULT_DOC_CATEGORIES = ["salary_slips", "tax", "insurance", "investments", "bank_statements"]

from werkzeug.utils import secure_filename


def _valid_category(name):
    """Check category name is safe (prevents path traversal)."""
    return bool(re.match(r'^[a-z][a-z0-9_]{0,49}$', name))


def _ensure_doc_dirs():
    """Create default documents folder structure on first run."""
    for cat in DEFAULT_DOC_CATEGORIES:
        for yr in range(2024, datetime.now().year + 1):
            (DOCS_DIR / cat / str(yr)).mkdir(parents=True, exist_ok=True)


@app.route("/api/documents/<category>/<year>")
def list_documents(category, year):
    """List all documents in a category/year folder."""
    if not _valid_category(category):
        return jsonify({"error": "Invalid category"}), 400
    if not re.match(r'^\d{4}$', year):
        return jsonify({"error": "Invalid year"}), 400

    folder = DOCS_DIR / category / year
    if not folder.exists():
        return jsonify([])

    files = []
    for f in sorted(folder.iterdir()):
        if f.is_file() and not f.name.startswith('.'):
            stat = f.stat()
            files.append({
                "name": f.name,
                "size": stat.st_size,
                "modified": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M"),
            })
    return jsonify(files)


@app.route("/api/documents/<category>/<year>/upload", methods=["POST"])
def upload_document(category, year):
    """Upload one or more files to a category/year folder."""
    if not _valid_category(category):
        return jsonify({"error": "Invalid category"}), 400
    if not re.match(r'^\d{4}$', year):
        return jsonify({"error": "Invalid year"}), 400

    folder = DOCS_DIR / category / year
    folder.mkdir(parents=True, exist_ok=True)

    uploaded = []
    for f in request.files.getlist("files"):
        if not f.filename:
            continue
        new_name = secure_filename(f.filename)
        stem = Path(new_name).stem
        ext = Path(new_name).suffix
        dest = folder / new_name
        # Avoid overwrite — append counter
        counter = 1
        while dest.exists():
            new_name = f"{stem}_{counter}{ext}"
            dest = folder / new_name
            counter += 1
        f.save(str(dest))
        uploaded.append(new_name)

    return jsonify({"ok": True, "uploaded": uploaded, "count": len(uploaded)})


@app.route("/api/documents/<category>/<year>/<filename>")
def download_document(category, year, filename):
    """Download a specific document."""
    if not _valid_category(category):
        return jsonify({"error": "Invalid category"}), 400
    if not re.match(r'^\d{4}$', year):
        return jsonify({"error": "Invalid year"}), 400
    safe_name = secure_filename(filename)
    folder = DOCS_DIR / category / year
    file_path = folder / safe_name
    if not file_path.exists() or not file_path.is_file():
        return jsonify({"error": "File not found"}), 404
    return send_from_directory(str(folder), safe_name,
                               as_attachment=request.args.get("download") is not None)


@app.route("/api/documents/<category>/<year>/<filename>", methods=["DELETE"])
def delete_document(category, year, filename):
    """Delete a specific document."""
    if not _valid_category(category):
        return jsonify({"error": "Invalid category"}), 400
    if not re.match(r'^\d{4}$', year):
        return jsonify({"error": "Invalid year"}), 400
    safe_name = secure_filename(filename)
    folder = DOCS_DIR / category / year
    file_path = folder / safe_name
    if not file_path.exists() or not file_path.is_file():
        return jsonify({"error": "File not found"}), 404
    file_path.unlink()
    return jsonify({"ok": True, "deleted": safe_name})


@app.route("/api/documents/categories")
def doc_categories():
    """Return available categories and years (discovered from disk)."""
    cats = {}
    if DOCS_DIR.exists():
        for d in sorted(DOCS_DIR.iterdir()):
            if d.is_dir() and not d.name.startswith('.') and _valid_category(d.name):
                years = sorted([y.name for y in d.iterdir() if y.is_dir() and y.name.isdigit()], reverse=True)
                cats[d.name] = years
    return jsonify(cats)


@app.route("/api/documents/categories", methods=["POST"])
def create_doc_category():
    """Create a new document category folder."""
    data = request.get_json(silent=True) or {}
    raw = data.get("name", "").strip().lower().replace(" ", "_")
    raw = re.sub(r'[^a-z0-9_]', '', raw)
    if not _valid_category(raw):
        return jsonify({"error": "Invalid name. Use letters, numbers, underscores."}), 400
    cat_dir = DOCS_DIR / raw
    if cat_dir.exists():
        return jsonify({"error": "Category already exists"}), 409
    (cat_dir / str(datetime.now().year)).mkdir(parents=True, exist_ok=True)
    return jsonify({"ok": True, "category": raw})


@app.route("/api/documents/categories/<category>", methods=["DELETE"])
def delete_doc_category(category):
    """Delete an empty document category."""
    if not _valid_category(category):
        return jsonify({"error": "Invalid category"}), 400
    cat_dir = DOCS_DIR / category
    if not cat_dir.exists():
        return jsonify({"error": "Category not found"}), 404
    for root, dirs, files in os.walk(str(cat_dir)):
        if files:
            return jsonify({"error": "Category is not empty. Delete all files first."}), 400
    shutil.rmtree(str(cat_dir))
    return jsonify({"ok": True, "deleted": category})


@app.route("/api/documents/categories/<category>/years", methods=["POST"])
def create_doc_year(category):
    """Create a year folder inside a category."""
    if not _valid_category(category):
        return jsonify({"error": "Invalid category"}), 400
    cat_dir = DOCS_DIR / category
    if not cat_dir.is_dir():
        return jsonify({"error": "Category not found"}), 404
    data = request.get_json(silent=True) or {}
    year = str(data.get("year", "")).strip()
    if not re.match(r'^\d{4}$', year):
        return jsonify({"error": "Invalid year"}), 400
    yr_dir = cat_dir / year
    if yr_dir.exists():
        return jsonify({"error": "Year already exists"}), 409
    yr_dir.mkdir(parents=True, exist_ok=True)
    return jsonify({"ok": True, "category": category, "year": year})


# ═══════════════════════════════════════════════════════════════
#  STARTUP
# ═══════════════════════════════════════════════════════════════

if __name__ == "__main__":
    _ensure_workbook()
    _ensure_doc_dirs()
    synced, sync_errors = _sync_form_expenses()
    if synced:
        print(f"[OK] Imported {synced} new expense(s) from Google Sheets")
    elif sync_errors:
        print(f"[!] Google Sheets expense sync skipped: {sync_errors[0]}")
    port = int(os.environ.get("PORT", 5000))
    host = os.environ.get("FINTRACK_HOST", "127.0.0.1")
    debug = os.environ.get("FLASK_DEBUG", "false").lower() in ("1", "true", "yes")
    print(f"\n  FinTrack server running at http://localhost:{port}\n")
    app.run(host=host, port=port, debug=debug, use_reloader=False)
